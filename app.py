import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import re
from datetime import datetime

# Current year for validation - prevent future years from being accepted
CURRENT_YEAR = datetime.now().year

# Optional OCR imports - app works without these
OCR_AVAILABLE = False
try:
    import pytesseract
    from PIL import Image
    from pdf2image import convert_from_path
    import tempfile
    import os
    OCR_AVAILABLE = True
    # Configure Tesseract path for Windows (adjust if installed elsewhere)
    # pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
except ImportError:
    pass  # OCR features will be disabled

def extract_tables_from_pdf(pdf_file):
    """Extract tables from all pages of the PDF."""
    all_tables = []
    
    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            for table in tables:
                if table:
                    # Handle duplicate column names by making them unique
                    headers = table[0]
                    unique_headers = []
                    header_counts = {}
                    for h in headers:
                        h_str = str(h) if h is not None else ''
                        if h_str in header_counts:
                            header_counts[h_str] += 1
                            h_str = f"{h_str}_{header_counts[h_str]}"
                        else:
                            header_counts[h_str] = 0
                        unique_headers.append(h_str)
                    df = pd.DataFrame(table[1:], columns=unique_headers)
                    all_tables.append({
                        'page': page_num,
                        'data': df
                    })
    
    return all_tables

def extract_text_to_dataframe(pdf_file):
    """Extract text from PDF and convert to DataFrame."""
    all_text = []
    
    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                for line in lines:
                    if line.strip():
                        all_text.append({
                            'Page': page_num,
                            'Content': line.strip()
                        })
    
    return pd.DataFrame(all_text)

def extract_tables_as_transactions(pdf_file):
    """
    First attempt: Try to extract tables and map columns intelligently.
    Bank statements often have tabular structure we can use.
    """
    transactions = []
    
    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                headers = [str(h).strip().lower() if h else '' for h in table[0]]
                
                # Map column indices based on headers
                date_idx = None
                desc_idx = None
                debit_idx = None
                credit_idx = None
                balance_idx = None
                amount_idx = None
                
                # Try to identify columns from headers
                for i, h in enumerate(headers):
                    if any(word in h for word in ['date', 'posted', 'value']):
                        date_idx = i
                    elif any(word in h for word in ['description', 'details', 'narrative', 'particulars', 'transaction']):
                        desc_idx = i
                    elif any(word in h for word in ['debit', 'withdrawal', 'dr', 'out']):
                        debit_idx = i
                    elif any(word in h for word in ['credit', 'deposit', 'cr', 'in']):
                        credit_idx = i
                    elif any(word in h for word in ['balance', 'bal', 'running']):
                        balance_idx = i
                    elif any(word in h for word in ['amount', 'value', 'money']):
                        amount_idx = i
                
                # If no headers found, try to infer from first data row
                if date_idx is None and len(table) > 1:
                    for i, cell in enumerate(table[1]):
                        if cell and re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}(\s+\d{4})?', str(cell), re.IGNORECASE):
                            date_idx = i
                            break
                
                # Process data rows
                for row in table[1:]:
                    if not row or len(row) < 2:
                        continue
                    
                    # Skip rows without dates (likely headers or summaries)
                    date_val = row[date_idx] if date_idx is not None and date_idx < len(row) else None
                    if not date_val or not str(date_val).strip():
                        continue
                    
                    # Skip if date doesn't look like a date
                    if not re.search(r'\d', str(date_val)):
                        continue
                    
                    # Extract description
                    description = ''
                    if desc_idx is not None and desc_idx < len(row):
                        description = str(row[desc_idx]).strip()
                    else:
                        # Use non-date, non-amount columns as description
                        desc_parts = []
                        for i, cell in enumerate(row):
                            if cell and i not in [date_idx, debit_idx, credit_idx, balance_idx, amount_idx]:
                                cell_str = str(cell).strip()
                                # Check if this is a number that might have lost leading zeros
                                # If it's a 6-digit number starting with 0 or 00, preserve as text
                                if cell_str and re.match(r'^\d{5,6}$', cell_str):
                                    # Check if original cell was a number (pdfplumber converted it)
                                    # If so, we can't recover leading zeros, but we can format it
                                    # For now, keep it as is since we can't recover original
                                    desc_parts.append(cell_str)
                                elif cell_str and not re.match(r'^[\d,\.\-\$€£¥]+$', cell_str):
                                    # Filter out "None" and "blank" strings from empty cells
                                    if cell_str.lower() not in ['none', 'blank']:
                                        desc_parts.append(cell_str)
                        description = ' '.join(desc_parts)
                    
                    # Extract amounts
                    debit = 0.0
                    credit = 0.0
                    balance = 0.0
                    
                    if debit_idx is not None and debit_idx < len(row):
                        val = str(row[debit_idx]).replace(',', '').replace('$', '').strip()
                        if val and re.match(r'^-?[\d\.]+$', val):
                            debit = abs(float(val))
                    
                    if credit_idx is not None and credit_idx < len(row):
                        val = str(row[credit_idx]).replace(',', '').replace('$', '').strip()
                        if val and re.match(r'^-?[\d\.]+$', val):
                            credit = abs(float(val))
                    
                    if balance_idx is not None and balance_idx < len(row):
                        val = str(row[balance_idx]).replace(',', '').replace('$', '').strip()
                        if val and re.match(r'^-?[\d\.]+$', val):
                            balance = float(val)
                    
                    # If we have an amount column but no separate debit/credit
                    if amount_idx is not None and amount_idx < len(row) and debit == 0 and credit == 0:
                        val = str(row[amount_idx]).replace(',', '').replace('$', '').strip()
                        if val:
                            # Check for negative sign or brackets indicating debit
                            is_negative = '-' in val or val.startswith('(') or 'dr' in val.lower()
                            num_val = re.sub(r'[^\d\.]', '', val)
                            if num_val:
                                if is_negative:
                                    debit = abs(float(num_val))
                                else:
                                    credit = abs(float(num_val))
                    
                    # If still no amounts, try to find them in any column
                    if debit == 0 and credit == 0:
                        for i, cell in enumerate(row):
                            if i == date_idx:
                                continue
                            val = str(cell).replace(',', '').strip()
                            nums = re.findall(r'-?[\d,]+\.\d{2}', val)
                            for num in nums:
                                num_f = float(num.replace(',', ''))
                                if num.startswith('-') or '(' in val:
                                    if abs(num_f) > debit:
                                        debit = abs(num_f)
                                else:
                                    if num_f > credit:
                                        credit = num_f
                    
                    transactions.append({
                        'Date': str(date_val).strip(),
                        'Transaction Description': description,
                        'Debit': debit,
                        'Credit': credit,
                        'Balance': balance
                    })
    
    return pd.DataFrame(transactions)

def parse_bank_statement_transactions(pdf_file):
    """
    Smart bank statement parser that tries table extraction first,
    then falls back to line-by-line parsing with better heuristics.
    """
    # First try table extraction
    df = extract_tables_as_transactions(pdf_file)
    if not df.empty and len(df) > 0:
        return df
    
    # Fallback: line-by-line parsing
    transactions = []
    last_balance = None
    
    # Year tracking for date inference
    current_year = None
    month_map = get_month_map()
    
    # Track last date on current page for chronological ordering
    last_date_on_page = None
    
    # Track whether we've found the first transaction yet
    # Only check for year markers before the first transaction is found
    first_transaction_found = False
    
    # Date patterns
    # Hard-coded month names to prevent random words from being matched
    month_names = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
    month_abbr = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
    date_patterns = [
        (r'((0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](20\d{2}|\d{2}))', 'numeric'),  # DD/MM/YYYY
        (rf'(\d{{1,2}})\s+{month_abbr}\s+(20\d{{2}}|\d{{2}})', 'text_abbr'),  # 17 Mar 2022 - moved before full month names
        (rf'(\d{{1,2}})\s+{month_names}\s+(20\d{{2}}|\d{{2}})', 'text'),  # 15 January 2024
        (r'(\d{1,2})(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)', 'compact'),  # 01JAN
    ]
    
    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if not text:
                continue
            
            # Reset last date tracker for new page
            last_date_on_page = None
            
            lines = text.split('\n')
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Check for year markers first (like "2022", "2023")
                # Only check for year markers before the first transaction is found
                # This prevents dates in transactions from being incorrectly flagged as year markers
                if not first_transaction_found:
                    # Check for StatementPeriod pattern (e.g., "StatementPeriod 30/12/2023to29/01/2024" or "Statement Period 30/12/2023 to 29/01/2024")
                    statement_period_match = re.match(r'Statement\s?Period\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*to\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line, re.IGNORECASE)
                    if statement_period_match:
                        date1_str = statement_period_match.group(1)
                        date2_str = statement_period_match.group(2)
                        # Try to extract years from both dates
                        year1 = None
                        year2 = None
                        year1_match = re.search(r'(\d{2,4})$', date1_str)
                        year2_match = re.search(r'(\d{2,4})$', date2_str)
                        if year1_match:
                            year1 = int(year1_match.group(1))
                            if year1 < 100:
                                year1 = 2000 + year1
                        if year2_match:
                            year2 = int(year2_match.group(1))
                            if year2 < 100:
                                year2 = 2000 + year2
                        # Use the first date's year as the statement year
                        if year1 and year1 >= 2000 and year1 <= CURRENT_YEAR:
                            current_year = year1
                        continue
                    
                    year_match = re.match(r'^\s*(20\d{2})\s*$', line)
                    year_start_match = re.match(r'^\s*(20\d{2})\b', line)
                    # Only check for year anywhere if line is very short (<= 8 chars) and doesn't look like a date
                    year_anywhere_match = None
                    if len(line) <= 8 and not re.search(r'\d{1,2}[/-]', line) and not re.search(r'[A-Za-z]{3}', line):
                        year_anywhere_match = re.search(r'\b(20\d{2})\b', line)
                    
                    if year_match or year_start_match or year_anywhere_match:
                        if year_match:
                            year_str = year_match.group(1)
                        elif year_start_match:
                            year_str = year_start_match.group(1)
                        else:
                            year_str = year_anywhere_match.group(1)
                        detected_year = int(year_str)
                        # Reject future years (e.g., 2077 from store numbers) and years before 2000
                        # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                        if detected_year > CURRENT_YEAR or detected_year < 2000:
                            if current_year is None:
                                current_year = CURRENT_YEAR
                            # Don't update current_year with the invalid year
                        else:
                            current_year = detected_year
                        continue
                
                # Check for date at start of line
                raw_date_str = None
                date_str = None  # Initialize to prevent UnboundLocalError
                for pattern, ptype in date_patterns:
                    match = re.match(pattern, line, re.IGNORECASE)
                    if match:
                        if ptype == 'numeric':
                            raw_date_str = match.group(0)
                        elif ptype == 'compact':
                            raw_date_str = match.group(0)
                        else:
                            raw_date_str = match.group(0)
                        break
                
                # Apply year inference and format to dd/mm/yyyy
                date_str, inferred_year = parse_date_with_year_inference(raw_date_str, current_year, month_map)
                if inferred_year:
                    # Reject inferred years before 2000 or after current year
                    # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                    if inferred_year < 2000 or inferred_year > CURRENT_YEAR:
                        if current_year is None:
                            current_year = CURRENT_YEAR
                        # Don't update current_year with the invalid year
                        # But still process the date with current_year
                        if date_str:
                            # Re-parse with valid current_year
                            date_str, _ = parse_date_with_year_inference(raw_date_str, current_year, month_map)
                    else:
                        current_year = inferred_year
                    # If this was just a year marker, skip it
                    if date_str is None:
                        continue
                    else:
                        # Chronological ordering: ensure date is not earlier than last date on page
                        if last_date_on_page and date_str:
                            try:
                                from datetime import datetime
                                current_date = datetime.strptime(date_str, '%d/%m/%Y')
                                last_date = datetime.strptime(last_date_on_page, '%d/%m/%Y')
                                if current_date < last_date:
                                    # Date is earlier, increment year
                                    new_year = current_date.year + 1
                                    if new_year <= CURRENT_YEAR:
                                        date_str = current_date.replace(year=new_year).strftime('%d/%m/%Y')
                                        current_year = new_year
                            except:
                                pass  # If date parsing fails, keep as is
                        # Update last date on page
                        last_date_on_page = date_str
                
                if not date_str:
                    continue
                
                # Find all amounts with their positions (handles +$100, -$100, $100, $100.00, etc.)
                amount_matches = list(re.finditer(r'[\+\-]?[\$€£¥]?\s*[\d,]+(?:\.\d{2})?', line))
                amounts = []
                for m in amount_matches:
                    amt_str = m.group().replace(',', '').replace('$', '').replace('€', '').replace('£', '').strip()
                    try:
                        amt_val = float(amt_str)
                        amounts.append({
                            'value': abs(amt_val),
                            'pos': m.start(),
                            'is_negative': m.group().strip().startswith('-') or amt_val < 0,
                            'is_positive': m.group().strip().startswith('+'),
                            'raw': m.group()
                        })
                    except:
                        continue
                
                if len(amounts) == 0:
                    continue
                
                # Extract description - everything between date and first amount
                date_end = len(raw_date_str) if raw_date_str else 0
                if amounts:
                    first_amt_pos = amounts[0]['pos']
                    description = line[date_end:first_amt_pos].strip()
                else:
                    description = line[date_end:].strip()
                
                # Clean description
                description = re.sub(r'^[\s\-\|\.]+', '', description).strip()
                description = re.sub(r'\s+', ' ', description)
                
                # Smart amount categorization
                debit = 0.0
                credit = 0.0
                balance = 0.0
                
                if len(amounts) == 1:
                    # Single amount - could be transaction or balance
                    # If we have a previous balance, compare to determine if it's a new balance
                    amt = amounts[0]
                    if last_balance is not None and abs(amt['value'] - last_balance) < 0.01:
                        # Same as last balance, might be just a reference
                        balance = amt['value']
                    elif amt.get('is_positive'):
                        credit = amt['value']
                    elif amt['is_negative'] or any(x in line[amounts[0]['pos']-5:amounts[0]['pos']].lower() for x in ['-', 'dr', 'debit']):
                        debit = amt['value']
                    elif any(x in line[amounts[0]['pos']-5:amounts[0]['pos']].lower() for x in ['+', 'cr', 'credit', 'deposit']):
                        credit = amt['value']
                    elif any(x in line.lower() for x in ['cr', 'credit', 'deposit']):
                        credit = amt['value']
                    else:
                        # Default: if amount is significant, treat as debit if no other indicators
                        debit = amt['value']
                    
                    if balance == 0:
                        balance = amt['value']
                        
                elif len(amounts) == 2:
                    # Two amounts - usually (transaction, balance) or (debit/credit, balance)
                    amt1, amt2 = amounts[0], amounts[1]
                    
                    # The larger position value (further right) is usually balance
                    if amt1['pos'] > amt2['pos']:
                        amt1, amt2 = amt2, amt1
                    
                    balance = amt2['value']
                    
                    # Determine if amt1 is debit or credit
                    if amt1.get('is_positive'):
                        credit = amt1['value']
                    elif amt1['is_negative']:
                        debit = amt1['value']
                    elif any(x in line[amt1['pos']-5:amt1['pos']].lower() for x in ['+', 'cr', 'credit']):
                        credit = amt1['value']
                    elif any(x in description.lower() for x in ['transfer from', 'deposit', 'credit', 'salary', 'refund']):
                        credit = amt1['value']
                    elif any(x in description.lower() for x in ['transfer to', 'withdrawal', 'payment', 'purchase', 'fee']):
                        debit = amt1['value']
                    else:
                        # Use balance change to determine
                        if last_balance is not None:
                            expected_new_balance = last_balance - amt1['value']
                            expected_new_balance_credit = last_balance + amt1['value']
                            if abs(expected_new_balance - balance) < 0.01:
                                debit = amt1['value']
                            elif abs(expected_new_balance_credit - balance) < 0.01:
                                credit = amt1['value']
                            else:
                                debit = amt1['value']  # Default to debit
                        else:
                            debit = amt1['value']
                            
                elif len(amounts) >= 3:
                    # Three columns: debit, credit, balance (common format)
                    # Sort by position
                    sorted_amts = sorted(amounts, key=lambda x: x['pos'])
                    
                    # Most bank statements: debit column first, then credit, then balance
                    if len(sorted_amts) >= 3:
                        # Check if middle value is 0 or very small (might be empty credit column)
                        if sorted_amts[1]['value'] < 0.01:
                            # Format: [amount] [0] [balance] - single column for both
                            if sorted_amts[0]['is_negative']:
                                debit = sorted_amts[0]['value']
                            else:
                                credit = sorted_amts[0]['value']
                            balance = sorted_amts[2]['value']
                        else:
                            # Format: [debit] [credit] [balance]
                            debit = sorted_amts[0]['value']
                            credit = sorted_amts[1]['value']
                            balance = sorted_amts[2]['value']
                
                # Update last balance for next iteration
                if balance > 0:
                    last_balance = balance
                
                # Only add if we have meaningful data
                if description or debit > 0 or credit > 0 or balance > 0:
                    first_transaction_found = True
                    transactions.append({
                        'Date': date_match,
                        'Transaction Description': description,
                        'Debit': round(debit, 2),
                        'Credit': round(credit, 2),
                        'Balance': round(balance, 2)
                    })
    
    return pd.DataFrame(transactions)

def create_excel_with_tables(tables, text_df):
    """Create Excel file from tables and text data."""
    output = io.BytesIO()
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Add sheets with tables
    if tables:
        for i, table_info in enumerate(tables):
            sheet_name = f"Table_{i+1}_Page_{table_info['page']}"
            # Truncate sheet name if too long (Excel limit: 31 chars)
            sheet_name = sheet_name[:31]
            
            ws = workbook.create_sheet(title=sheet_name)
            df = table_info['data']
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add text content sheet
    if not text_df.empty:
        ws_text = workbook.create_sheet(title="Text_Content")
        for r_idx, row in enumerate(dataframe_to_rows(text_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_text.cell(row=r_idx, column=c_idx, value=value)
    
    # If no data was extracted, add a placeholder sheet
    if len(workbook.sheetnames) == 0:
        ws = workbook.create_sheet(title="No_Data")
        ws['A1'] = "No data could be extracted from this PDF."
    
    workbook.save(output)
    output.seek(0)
    
    return output

def find_amount_columns(lines_with_amounts):
    """
    Analyze the physical positions of amounts to identify column boundaries.
    Returns: (debit_column_range, credit_column_range, balance_column_range)
    """
    if not lines_with_amounts:
        return None, None, None
    
    # Collect all amount positions
    all_positions = []
    for line_data in lines_with_amounts:
        for amt in line_data['amounts']:
            all_positions.append({
                'start': amt['start'],
                'end': amt['end'],
                'value': amt['value'],
                'line': line_data['line']
            })
    
    if not all_positions:
        return None, None, None
    
    # Cluster positions to find columns
    # Sort by position
    all_positions.sort(key=lambda x: x['start'])
    
    # Simple clustering: group positions that are close together
    clusters = []
    current_cluster = [all_positions[0]]
    
    for pos in all_positions[1:]:
        # If within 10 chars of cluster average, add to cluster
        cluster_avg = sum(p['start'] for p in current_cluster) / len(current_cluster)
        if abs(pos['start'] - cluster_avg) < 15:
            current_cluster.append(pos)
        else:
            clusters.append(current_cluster)
            current_cluster = [pos]
    
    if current_cluster:
        clusters.append(current_cluster)
    
    # Map clusters to columns (rightmost is usually balance)
    clusters.sort(key=lambda c: sum(p['start'] for p in c) / len(c))
    
    column_ranges = []
    for cluster in clusters:
        starts = [p['start'] for p in cluster]
        ends = [p['end'] for p in cluster]
        column_ranges.append({
            'min_start': min(starts),
            'max_end': max(ends),
            'avg_start': sum(starts) / len(starts),
            'count': len(cluster)
        })
    
    # Assign columns based on typical bank statement layout
    # Usually: [description area] [debit] [credit] [balance]
    # Or: [description] [amount] [balance] where amount can be debit or credit
    
    if len(column_ranges) >= 3:
        # Likely have debit, credit, balance columns
        return column_ranges[-3], column_ranges[-2], column_ranges[-1]
    elif len(column_ranges) == 2:
        # Likely have amount and balance
        return column_ranges[-2], None, column_ranges[-1]
    elif len(column_ranges) == 1:
        # Just balance or single amount column
        return None, None, column_ranges[-1]
    
    return None, None, None


def is_scanned_pdf(pdf_file):
    """
    Check if a PDF appears to be scanned (image-based) rather than text-based.
    Returns True if the PDF has no extractable text on most pages.
    """
    text_count = 0
    image_count = 0
    
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages[:3]:  # Check first 3 pages
            text = page.extract_text()
            if text and len(text.strip()) > 50:  # Has substantial text
                text_count += 1
            else:
                image_count += 1
    
    # If more than half the checked pages lack text, assume it's scanned
    return image_count > text_count


def ocr_image_to_text(image):
    """
    Run OCR on a PIL Image and return extracted text.
    Uses Tesseract with configuration optimized for documents.
    """
    if not OCR_AVAILABLE:
        st.error("❌ OCR not available. Please install pytesseract and Tesseract OCR.")
        return ""
    
    # Configuration for better accuracy on financial documents
    # --psm 6 = Assume a single uniform block of text
    # -c tessedit_char_whitelist = Limit characters to common ones
    custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist="0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,/-$€£¥%&()[]:; "'
    
    try:
        text = pytesseract.image_to_string(image, config=custom_config)
        return text.strip()
    except Exception as e:
        st.warning(f"OCR failed: {str(e)}")
        return ""


def ocr_pdf_page(page_image):
    """
    OCR a single PDF page (provided as PIL Image) and return structured text.
    Also attempts to extract word positions for coordinate-based parsing.
    """
    if not OCR_AVAILABLE:
        st.error("❌ OCR not available. Please install pytesseract and Tesseract OCR.")
        return {'text': '', 'words': [], 'lines': []}
    
    # Get OCR data with bounding boxes
    try:
        data = pytesseract.image_to_data(page_image, output_type=pytesseract.Output.DICT)
        
        words = []
        lines = []
        current_line_y = None
        current_line_words = []
        
        for i, text in enumerate(data['text']):
            if int(data['conf'][i]) > 30 and text.strip():  # Filter low confidence
                word_info = {
                    'text': text,
                    'x0': data['left'][i],
                    'x1': data['left'][i] + data['width'][i],
                    'y0': data['top'][i],
                    'y1': data['top'][i] + data['height'][i],
                    'confidence': data['conf'][i]
                }
                words.append(word_info)
                
                # Group into lines by y-position (with tolerance)
                y_center = (word_info['y0'] + word_info['y1']) / 2
                if current_line_y is None or abs(y_center - current_line_y) > 10:
                    if current_line_words:
                        lines.append(current_line_words)
                    current_line_words = [word_info]
                    current_line_y = y_center
                else:
                    current_line_words.append(word_info)
        
        if current_line_words:
            lines.append(current_line_words)
        
        # Build text representation
        full_text = '\n'.join(' '.join(w['text'] for w in line) for line in lines)
        
        return {
            'text': full_text,
            'words': words,
            'lines': lines
        }
    except Exception as e:
        st.warning(f"OCR with position data failed: {str(e)}")
        # Fallback to simple text OCR
        text = pytesseract.image_to_string(page_image)
        return {'text': text, 'words': [], 'lines': []}


def process_pdf_with_ocr(pdf_file, page_filter=None):
    """
    Process a scanned PDF using OCR. Converts pages to images and runs OCR.
    Returns structured data compatible with the visual parsing functions.
    """
    if not OCR_AVAILABLE:
        st.error("❌ OCR not available. Please install pytesseract and Tesseract OCR.")
        return []
    
    all_pages_data = []
    
    # Save PDF to temp file for pdf2image
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
        pdf_file.seek(0)
        tmp_pdf.write(pdf_file.read())
        tmp_path = tmp_pdf.name
    
    try:
        # Convert PDF pages to images
        images = convert_from_path(tmp_path, dpi=300)  # High DPI for better accuracy
        
        for page_num, image in enumerate(images, 1):
            if page_filter and page_num not in page_filter:
                continue
                
            with st.spinner(f"Running OCR on page {page_num}..."):
                ocr_result = ocr_pdf_page(image)
                all_pages_data.append({
                    'page_num': page_num,
                    'text': ocr_result['text'],
                    'words': ocr_result['words'],
                    'lines': ocr_result['lines']
                })
    finally:
        os.unlink(tmp_path)
    
    return all_pages_data


def parse_date_with_year_inference(date_str, current_year, month_map):
    """
    Parse a date string and return it in dd/mm/yyyy format.
    Handles dates with and without years, using year inference.
    
    Args:
        date_str: The date string from the PDF (e.g., "31MAY", "01/08/2022", "29JUL")
        current_year: The most recently detected year marker (e.g., 2022)
        month_map: Dict mapping month abbreviations to numbers
    
    Returns:
        Tuple of (formatted_date_str, year_if_found)
        formatted_date_str is in dd/mm/yyyy format
    """
    if not date_str:
        return '', None
    
    date_str = date_str.strip().upper()
    
    # Check if this is just a year (e.g., "2022", "2023")
    year_only_match = re.match(r'^(20\d{2})\s*$', date_str)
    if year_only_match:
        year = int(year_only_match.group(1))
        # Reject future years and years before 2000
        # Instead of returning None, use current_year if available
        if year > CURRENT_YEAR or year < 2000:
            if current_year:
                return None, current_year
            return None, CURRENT_YEAR
        return None, year
    
    # Pattern 1: Numeric date with 2-digit year (DD/MM/YY or DD-MM-YY) - interpret as 20xx
    numeric_2digit = re.match(r'((0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](\d{2}))$', date_str)
    if numeric_2digit:
        _, d, m, y = numeric_2digit.groups()
        year_2digit = int(y)
        # Interpret 2-digit year as 20xx (2000-2099)
        year = 2000 + year_2digit
        # Reject future years
        if year > CURRENT_YEAR:
            if current_year:
                return f"{int(d):02d}/{int(m):02d}/{current_year}", current_year
            return f"{int(d):02d}/{int(m):02d}/{CURRENT_YEAR}", CURRENT_YEAR
        return f"{int(d):02d}/{int(m):02d}/{year}", year
    
    # Pattern 2: Numeric date with 4-digit year (DD/MM/YYYY or DD-MM-YYYY)
    numeric_match = re.match(r'((0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](\d{4}))', date_str)
    if numeric_match:
        _, d, m, y = numeric_match.groups()
        year = int(y)
        # Reject future years and years before 2000
        # Instead of returning None, use current_year if available
        if year > CURRENT_YEAR or year < 2000:
            if current_year:
                return f"{int(d):02d}/{int(m):02d}/{current_year}", current_year
            return f"{int(d):02d}/{int(m):02d}/{CURRENT_YEAR}", CURRENT_YEAR
        return f"{int(d):02d}/{int(m):02d}/{y}", year
    
    # Pattern 2: Compact format with year (DDMONYYYY like 01AUG2022)
    # Hard-coded month abbreviations to prevent random words from being matched
    compact_with_year = re.match(r'(\d{1,2})(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)(\d{4})', date_str)
    if compact_with_year:
        d, m, y = compact_with_year.groups()
        year = int(y)
        # Reject future years and years before 2000
        # Instead of returning None, use current_year if available
        if year > CURRENT_YEAR or year < 2000:
            month_num = month_map.get(m, 0)
            if month_num:
                if current_year:
                    return f"{int(d):02d}/{month_num:02d}/{current_year}", current_year
                return f"{int(d):02d}/{month_num:02d}/{CURRENT_YEAR}", CURRENT_YEAR
        month_num = month_map.get(m, 0)
        if month_num:
            return f"{int(d):02d}/{month_num:02d}/{y}", year
    
    # Pattern 3: Compact format without year (DDMON like 31MAY)
    # Hard-coded month abbreviations to prevent random words from being matched
    compact_no_year = re.match(r'(\d{1,2})(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)', date_str)
    if compact_no_year:
        d, m = compact_no_year.groups()
        month_num = month_map.get(m, 0)
        if month_num and current_year:
            return f"{int(d):02d}/{month_num:02d}/{current_year}", None
        elif month_num:
            # No year available yet - return partial date
            return f"{int(d):02d}/{month_num:02d}/", None
    
    # Pattern 4: Text format with year (DD Month YYYY like "15 August 2024")
    # Hard-coded month names to prevent random words from being matched
    text_with_year = re.match(r'(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', date_str)
    if text_with_year:
        d, m, y = text_with_year.groups()
        # Try to find month number
        month_num = month_map.get(m[:3].upper(), 0)
        if month_num:
            return f"{int(d):02d}/{month_num:02d}/{y}", int(y)
    
    # Pattern 4b: Text format with abbreviated month and year (DD Mon YYYY like "17 Mar 2022")
    # Hard-coded month abbreviations to prevent random words from being matched
    text_abbr_with_year = re.match(r'(\d{1,2})\s+(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(\d{4})', date_str)
    if text_abbr_with_year:
        d, m, y = text_abbr_with_year.groups()
        month_num = month_map.get(m, 0)
        if month_num:
            year = int(y)
            # Reject future years and years before 2000
            if year > CURRENT_YEAR or year < 2000:
                if current_year:
                    return f"{int(d):02d}/{month_num:02d}/{current_year}", current_year
                return f"{int(d):02d}/{month_num:02d}/{CURRENT_YEAR}", CURRENT_YEAR
            return f"{int(d):02d}/{month_num:02d}/{y}", year
    
    # Pattern 5: Text format without year (DD Month like "15 August")
    # Hard-coded month names to prevent random words from being matched
    text_no_year = re.match(r'(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)', date_str)
    if text_no_year:
        d, m = text_no_year.groups()
        month_num = month_map.get(m[:3].upper(), 0)
        if month_num and current_year:
            return f"{int(d):02d}/{month_num:02d}/{current_year}", None
        elif month_num:
            return f"{int(d):02d}/{month_num:02d}/", None
    
    # Pattern 5b: Text format with abbreviated month without year (DD Mon like "17 Mar")
    # Hard-coded month abbreviations to prevent random words from being matched
    text_abbr_no_year = re.match(r'(\d{1,2})\s+(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)', date_str)
    if text_abbr_no_year:
        d, m = text_abbr_no_year.groups()
        month_num = month_map.get(m, 0)
        if month_num and current_year:
            return f"{int(d):02d}/{month_num:02d}/{current_year}", None
        elif month_num:
            return f"{int(d):02d}/{month_num:02d}/", None
    
    # If no pattern matched, return original
    return date_str, None


def get_month_map():
    """Return mapping of month abbreviations to month numbers."""
    return {
        'JAN': 1, 'JANUARY': 1,
        'FEB': 2, 'FEBRUARY': 2,
        'MAR': 3, 'MARCH': 3,
        'APR': 4, 'APRIL': 4,
        'MAY': 5,
        'JUN': 6, 'JUNE': 6,
        'JUL': 7, 'JULY': 7,
        'AUG': 8, 'AUGUST': 8,
        'SEP': 9, 'SEPT': 9, 'SEPTEMBER': 9,
        'OCT': 10, 'OCTOBER': 10,
        'NOV': 11, 'NOVEMBER': 11,
        'DEC': 12, 'DECEMBER': 12
    }


def parse_page_spec(spec_str, total_pages):
    """
    Parse a page specification string into a set of page numbers.
    
    Supports formats like:
    - "1-9" -> pages 1 through 9
    - "1,2,3,10,25" -> pages 1, 2, 3, 10, and 25
    - "1-3,5,7-9" -> pages 1, 2, 3, 5, 7, 8, 9
    
    Args:
        spec_str: Page specification string
        total_pages: Total number of pages in PDF
    
    Returns: Set of page numbers (1-indexed)
    """
    if not spec_str or not spec_str.strip():
        return set(range(1, total_pages + 1))
    
    pages = set()
    parts = spec_str.replace(' ', '').split(',')
    
    for part in parts:
        if '-' in part:
            # Range like "1-9"
            try:
                start, end = part.split('-')
                start = int(start)
                end = int(end)
                for p in range(start, end + 1):
                    if 1 <= p <= total_pages:
                        pages.add(p)
            except ValueError:
                continue
        else:
            # Single page like "5"
            try:
                p = int(part)
                if 1 <= p <= total_pages:
                    pages.add(p)
            except ValueError:
                continue
    
    return pages


def parse_transactions_visual(pdf_file, debug_info=None, multiline_mode="Auto-detect", page_filter=None):
    """
    Parse transactions using VISUAL PDF coordinates (x, y positions).
    Detects column headers by their visual position on the page,
    then assigns amounts to columns based on their x-coordinate.
    
    Args:
        pdf_file: PDF file to parse
        debug_info: Optional dict to store debug information
        multiline_mode: How to handle multi-line descriptions
                       "Auto-detect" = smart detection
                       "Merge with PREVIOUS transaction" = always append to previous
                       "Merge with NEXT transaction" = always prepend to next
        page_filter: Optional set of page numbers to process (1-indexed), or None for all pages
    
    Returns: (transactions_df, all_content_df)
    """
    transactions = []
    all_content = []
    
    if debug_info is None:
        debug_info = {}
    
    debug_info['headers_found'] = []
    debug_info['column_boundaries'] = None
    debug_info['transactions_debug'] = []
    debug_info['year_determination_debug'] = []  # Track how years are determined
    debug_info['reconciliation'] = {
        'starting_balance': None,
        'running_balance': 0.0,
        'discrepancies': [],
        'cumulative_error': 0.0  # Track running total of discrepancies
    }
    
    # Date pattern (optional) - now also matches dates without year (e.g., "1 Jan", "15 Aug", "01JAN")
    # Made more strict to avoid matching random numbers in descriptions
    # Only accept 4-digit years starting with 20, or 2-digit years (which will be interpreted as 20xx)
    # Hard-coded month names to prevent random words from being matched
    month_names = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
    month_abbr = r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)'
    month_abbr_mixed = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
    date_pattern = rf'^((0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](20\d{{2}}|\d{{2}})|\d{{1,2}}\s+{month_names}(?:\s+20\d{{2}})?|\d{{1,2}}\s+{month_abbr_mixed}(?:\s+20\d{{2}})?|20\d{{4}}[/-](0?[1-9]|1[0-2])[/-](0?[1-9]|[12][0-9]|3[01])|\d{{1,2}}\s+{month_names}|\d{{1,2}}{month_abbr})'
    
    # Column header patterns for all 5 columns
    date_patterns = [r'\bdate\b', r'\b0date\b', r'\btransaction\s+date\b', r'\bvalue\s+date\b', r'\bposting\s+date\b']
    description_patterns = [r'\bdescription\b', r'\bparticulars\b', r'\bdetails\b', r'\btransaction\b', r'\bnarrative\b', r'\bmemo\b', r'\btransaction\s+details\b']
    debit_patterns = [r'\bdeb(it)?s?\b', r'\bwithdrawals?\b', r'\bpayments?\b', r'\bdr\b']
    credit_patterns = [r'\bcr(edits?)?\b', r'\bdeposits?\b', r'\breceived\b', r'\bcr\b']
    amount_patterns = [r'\bamount\b']
    balance_patterns = [r'\bbal(ance)?\b', r'\brunning\s+bal\b', r'\bavail\w*\s+bal\b']
    
    # Transaction section header indicators
    header_patterns = [
        r'transaction\s+description',
        r'date.*description',
        r'transactions?\s+details?',
        r'activity\s+details?',
    ]
    
    with pdfplumber.open(pdf_file) as pdf:
        column_boundaries = None
        in_transaction_section = False
        total_pages = len(pdf.pages)
        
        # Year tracking for date inference
        current_year = None
        month_map = get_month_map()
        
        # Track last date on current page for chronological ordering
        last_date_on_page = None
        
        # Track date range for the page (e.g., "3 nov 2022 to 3 feb 2023")
        date_range = None  # Will store {'start_date': datetime, 'end_date': datetime, 'start_year': int, 'end_year': int}
        
        # Apply page filter if provided
        if page_filter is None:
            page_filter = set(range(1, total_pages + 1))
        
        for page_num, page in enumerate(pdf.pages, 1):
            # Skip pages not in the filter
            if page_num not in page_filter:
                continue
            
            # Reset last date tracker for new page
            last_date_on_page = None
            # Reset date range for new page
            date_range = None
            
            # Statement period tracking (start and end dates)
            statement_start_date = None
            statement_end_date = None
            
            # Get text with visual positions (bbox coordinates)
            words = page.extract_words()
            if not words:
                continue
            
            # Group words by their y-position (same line)
            lines = group_words_by_line(words)
            
            # Check for StatementPeriod pattern at the very start of the page (before any header detection)
            # This should be done before column header detection to ensure we capture the year
            for line_words in lines[:5]:  # Check first 5 lines only
                line_text = ' '.join(w['text'] for w in line_words)
                line_stripped = line_text.strip()
                statement_period_match = re.match(r'Statement\s?Period\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*to\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line_stripped, re.IGNORECASE)
                if statement_period_match:
                    date1_str = statement_period_match.group(1)
                    date2_str = statement_period_match.group(2)
                    # Parse both dates as datetime objects for statement period tracking
                    try:
                        from datetime import datetime
                        # Parse first date
                        date1_parts = re.split(r'[/-]', date1_str)
                        if len(date1_parts) == 3:
                            d1, m1, y1 = date1_parts
                            if len(y1) == 2:
                                y1 = '20' + y1
                            statement_start_date = datetime(int(y1), int(m1), int(d1))
                        
                        # Parse second date
                        date2_parts = re.split(r'[/-]', date2_str)
                        if len(date2_parts) == 3:
                            d2, m2, y2 = date2_parts
                            if len(y2) == 2:
                                y2 = '20' + y2
                            statement_end_date = datetime(int(y2), int(m2), int(d2))
                        
                        debug_info['year_determination_debug'].append({
                            'page': page_num,
                            'line': line_stripped,
                            'type': 'STATEMENT_PERIOD_DETECTED',
                            'start_date': str(statement_start_date),
                            'end_date': str(statement_end_date)
                        })
                        all_content.append({
                            'Page': page_num,
                            'Content': f"[STATEMENT PERIOD] {line_stripped}"
                        })
                        break  # Only need to find it once per page
                    except Exception as e:
                        pass
            
            # At the start of each page, scan for column headers
            # First page: scan ALL lines (headers might be anywhere)
            # Subsequent pages: scan first 10 lines only
            if column_boundaries is None:
                lines_to_scan = lines if page_num == 1 else lines[:10]
                # Collect all potential header lines and choose the best one
                potential_headers = []
                for line_idx, line_words in enumerate(lines_to_scan):
                    test_boundaries, test_headers = detect_visual_column_boundaries_with_debug(
                        line_words, date_patterns, description_patterns, debit_patterns, credit_patterns, amount_patterns, balance_patterns
                    )
                    # Need at least 3 headers detected to be confident (Date, Desc, and one amount column)
                    if test_headers and len(test_headers) >= 3:
                        line_text = ' '.join(w['text'] for w in line_words)
                        line_lower = line_text.lower()
                        
                        # Skip lines that look like summary lines (opening, closing, total)
                        is_summary_line = any(kw in line_lower for kw in ['opening', 'closing', 'total', 'summary'])
                        
                        # Prioritize lines that contain 'date' header
                        has_date_header = any(h['type'] == 'DATE' for h in test_headers)
                        
                        # Score the line: +2 for date header, -1 for summary line
                        score = (2 if has_date_header else 0) - (1 if is_summary_line else 0)
                        
                        potential_headers.append({
                            'line_idx': line_idx,
                            'line_text': line_text,
                            'test_boundaries': test_boundaries,
                            'test_headers': test_headers,
                            'score': score,
                            'line_words': line_words
                        })
                
                # Choose the header line with the highest score
                if potential_headers:
                    best_header = max(potential_headers, key=lambda x: x['score'])
                    # Apply equalization to debit/credit ranges
                    column_boundaries = equalize_debit_credit_ranges(best_header['test_boundaries'])
                    header_info = best_header['test_headers']
                    line_text = best_header['line_text']
                    line_idx = best_header['line_idx']
                    line_words = best_header['line_words']
                    
                    # Debug: show what words were checked
                    words_checked = [{'text': w['text'], 'x': round((w['x0']+w['x1'])/2, 1)} for w in line_words]
                    
                    debug_info['headers_found'].append({
                        'page': page_num,
                        'line': line_text,
                        'headers': header_info,
                        'words_checked': words_checked
                    })
                    debug_info['column_boundaries'] = column_boundaries
                    
                    all_content.append({
                        'Page': page_num,
                        'Content': f"[COLUMN HEADER - Line {line_idx+1}] {line_text}"
                    })
                    in_transaction_section = True
            
            for line_idx, line_words in enumerate(lines):
                # Reconstruct line text
                line_text = ' '.join(w['text'] for w in line_words)
                line_stripped = line_text.strip()
                if not line_stripped or len(line_stripped) < 3:
                    continue
                
                # Debug: Show first few lines of each page
                if line_idx < 15 and page_num == 1:
                    st.write(f"🔍 Line {line_idx}: '{line_stripped}' | len={len(line_stripped)}")
                
                line_lower = line_stripped.lower()
                
                # Simple check: if we see keywords indicating transaction section start
                is_section_header = any(re.search(pattern, line_lower) for pattern in header_patterns)
                if is_section_header:
                    in_transaction_section = True
                    all_content.append({
                        'Page': page_num,
                        'Content': f"[SECTION HEADER] {line_stripped}"
                    })
                    continue
                
                # If we haven't found column headers yet, store as regular content
                if not in_transaction_section:
                    # Check for StatementPeriod pattern (e.g., "StatementPeriod 30/12/2023to29/01/2024" or "Statement Period 30/12/2023 to 29/01/2024")
                    statement_period_match = re.match(r'Statement\s?Period\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*to\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line_stripped, re.IGNORECASE)
                    if statement_period_match:
                        date1_str = statement_period_match.group(1)
                        date2_str = statement_period_match.group(2)
                        # Try to extract years from both dates
                        year1 = None
                        year2 = None
                        year1_match = re.search(r'(\d{2,4})$', date1_str)
                        year2_match = re.search(r'(\d{2,4})$', date2_str)
                        if year1_match:
                            year1 = int(year1_match.group(1))
                            if year1 < 100:
                                year1 = 2000 + year1
                        if year2_match:
                            year2 = int(year2_match.group(1))
                            if year2 < 100:
                                year2 = 2000 + year2
                        # Use the first date's year as the statement year
                        if year1 and year1 >= 2000 and year1 <= CURRENT_YEAR:
                            current_year = year1
                            debug_info['year_determination_debug'].append({
                                'page': page_num,
                                'line': line_stripped,
                                'type': 'STATEMENT_PERIOD_DETECTED',
                                'detected_year': current_year
                            })
                            all_content.append({
                                'Page': page_num,
                                'Content': f"[STATEMENT PERIOD] {line_stripped} (Year: {current_year})"
                            })
                            continue
                    
                    # Check for year markers only before section headers (like "2022", "2023")
                    # This prevents dates in transactions from being incorrectly flagged as year markers
                    # Check if line is just a year (allow leading/trailing whitespace)
                    year_match = re.match(r'^(20\d{2})\s*$', line_stripped)
                    # Check for year markers (numbers like 2023, 2024 at the start of line)
                    year_start_match = re.match(r'^(20\d{2})\s+', line_stripped)
                    # Find ALL years in the line (not just the first one)
                    year_all_matches = re.findall(r'\b(20\d{2})\b', line_stripped)
                    
                    # Check if this line contains a date range (two different years)
                    if len(year_all_matches) >= 2:
                        unique_years = list(set(year_all_matches))
                        if len(unique_years) >= 2:
                            # This looks like a date range summary (e.g., "3 nov 2022 to 3 feb 2023")
                            # Try to parse the dates in the line
                            try:
                                from datetime import datetime
                                # Find all date patterns in the line
                                date_pattern_with_year = r'(\d{1,2})\s+([A-Za-z]{3,9})\s+(20\d{2})'
                                date_matches = re.findall(date_pattern_with_year, line_stripped, re.IGNORECASE)
                                
                                if len(date_matches) >= 2:
                                    # Parse the dates
                                    dates = []
                                    for day, month, year in date_matches:
                                        try:
                                            month_num = month_map.get(month.upper(), 0)
                                            if month_num:
                                                dt = datetime(int(year), month_num, int(day))
                                                dates.append(dt)
                                        except:
                                            pass
                                    
                                    if len(dates) >= 2:
                                        # Sort dates to get start and end
                                        dates.sort()
                                        date_range = {
                                            'start_date': dates[0],
                                            'end_date': dates[-1],
                                            'start_year': dates[0].year,
                                            'end_year': dates[-1].year,
                                            'line': line_stripped
                                        }
                                        debug_info['year_determination_debug'].append({
                                            'page': page_num,
                                            'line': line_stripped,
                                            'type': 'DATE_RANGE_DETECTED',
                                            'start_date': dates[0].strftime('%d/%m/%Y'),
                                            'end_date': dates[-1].strftime('%d/%m/%Y'),
                                            'years': unique_years
                                        })
                                        # Don't treat this as a regular year marker
                                        year_match = None
                                        year_start_match = None
                                        year_all_matches = []
                            except:
                                pass
                    
                    if year_match or year_start_match or year_all_matches:
                        # Check if line has amounts before treating as year marker
                        # This prevents lines like "2023 25000.00" from being skipped
                        amounts_check = extract_amounts_with_positions(line_words)
                        if len(amounts_check) < 1:
                            # Additional check: exclude years that are part of monetary amounts
                            # Check if the year is preceded by currency symbol or decimal point
                            is_year_in_amount = False
                            for year_str in year_all_matches:
                                # Check if year is preceded by currency symbol or decimal
                                if re.search(r'[\$\€\£\¥\.\,]' + re.escape(year_str), line_stripped):
                                    is_year_in_amount = True
                                    break
                                # Check if year is followed by decimal point (part of larger number)
                                if re.search(re.escape(year_str) + r'\.\d+', line_stripped):
                                    is_year_in_amount = True
                                    break
                            
                            if is_year_in_amount:
                                debug_info['year_determination_debug'].append({
                                    'page': page_num,
                                    'line': line_stripped,
                                    'type': 'YEAR_MARKER_REJECTED',
                                    'detected_year': year_all_matches[0] if year_all_matches else 'N/A',
                                    'reason': 'Year is part of a monetary amount'
                                })
                            else:
                                if year_match:
                                    year_str = year_match.group(1)
                                elif year_start_match:
                                    year_str = year_start_match.group(1)
                                elif year_all_matches:
                                    # Use the first year found
                                    year_str = year_all_matches[0]
                                else:
                                    year_str = None  # Should not reach here
                                
                                if year_str:
                                    detected_year = int(year_str)
                                    # Reject future years (e.g., 2077 from store numbers) and years before 2000
                                    # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                                    if detected_year > CURRENT_YEAR or detected_year < 2000:
                                        if current_year is None:
                                            current_year = CURRENT_YEAR
                                        # Don't update current_year with the invalid year
                                        debug_info['year_determination_debug'].append({
                                            'page': page_num,
                                            'line': line_stripped,
                                            'type': 'YEAR_MARKER_REJECTED',
                                            'detected_year': detected_year,
                                            'reason': f'Year {detected_year} outside valid range (2000-{CURRENT_YEAR})',
                                            'current_year': current_year
                                        })
                                    else:
                                        current_year = detected_year
                                        debug_info['year_determination_debug'].append({
                                            'page': page_num,
                                            'line': line_stripped,
                                            'type': 'YEAR_MARKER_DETECTED',
                                            'detected_year': detected_year,
                                            'current_year_set': current_year
                                        })
                                        all_content.append({
                                            'Page': page_num,
                                            'Content': f"[YEAR MARKER] {current_year}"
                                        })
                                        continue
                    
                    all_content.append({
                        'Page': page_num,
                        'Content': line_stripped
                    })
                    continue
                
                # From here on, we're in the transaction section
                
                # Check for date at the beginning
                date_match = re.match(date_pattern, line_stripped)
                raw_date_str = date_match.group(1) if date_match else None
                
                # Initialize date_str to prevent UnboundLocalError
                date_str = None
                
                # Apply year inference and format to dd/mm/yyyy
                date_str, inferred_year = parse_date_with_year_inference(raw_date_str, current_year, month_map)
                
                # If statement period is available, use it to determine the correct year
                if statement_start_date and statement_end_date and date_str:
                    try:
                        from datetime import datetime
                        # Check if date_str has a complete year (format dd/mm/yyyy)
                        if date_str.count('/') == 2 and len(date_str.split('/')[-1]) == 4:
                            # Date already has a year, parse it
                            trans_date = datetime.strptime(date_str, '%d/%m/%Y')
                            
                            # Check if the transaction date falls within the statement period
                            # Try both years from the statement period
                            trans_date_start_year = trans_date.replace(year=statement_start_date.year)
                            trans_date_end_year = trans_date.replace(year=statement_end_date.year)
                            
                            # Determine which year is correct based on statement period
                            if statement_start_date <= trans_date_start_year <= statement_end_date:
                                # Transaction falls in statement period with start year
                                date_str = trans_date_start_year.strftime('%d/%m/%Y')
                                inferred_year = statement_start_date.year
                            elif statement_start_date <= trans_date_end_year <= statement_end_date:
                                # Transaction falls in statement period with end year
                                date_str = trans_date_end_year.strftime('%d/%m/%Y')
                                inferred_year = statement_end_date.year
                            else:
                                # Transaction doesn't fall in either year of statement period
                                # Use the year that's closer to the statement period
                                if abs(trans_date_start_year.year - statement_start_date.year) < abs(trans_date_end_year.year - statement_end_date.year):
                                    date_str = trans_date_start_year.strftime('%d/%m/%Y')
                                    inferred_year = statement_start_date.year
                                else:
                                    date_str = trans_date_end_year.strftime('%d/%m/%Y')
                                    inferred_year = statement_end_date.year
                        else:
                            # Date doesn't have a year yet (format dd/mm/), use statement period to assign year
                            # Extract day and month from date_str
                            parts = date_str.split('/')
                            if len(parts) >= 2:
                                day = int(parts[0])
                                month = int(parts[1])
                                
                                # Try both years from statement period
                                trans_date_start_year = datetime(statement_start_date.year, month, day)
                                trans_date_end_year = datetime(statement_end_date.year, month, day)
                                
                                # Determine which year is correct based on statement period
                                if statement_start_date <= trans_date_start_year <= statement_end_date:
                                    # Transaction falls in statement period with start year
                                    date_str = f"{day:02d}/{month:02d}/{statement_start_date.year}"
                                    inferred_year = statement_start_date.year
                                elif statement_start_date <= trans_date_end_year <= statement_end_date:
                                    # Transaction falls in statement period with end year
                                    date_str = f"{day:02d}/{month:02d}/{statement_end_date.year}"
                                    inferred_year = statement_end_date.year
                                else:
                                    # Transaction doesn't fall in either year of statement period
                                    # Use the year that's closer to the statement period
                                    if abs(trans_date_start_year.year - statement_start_date.year) < abs(trans_date_end_year.year - statement_end_date.year):
                                        date_str = f"{day:02d}/{month:02d}/{statement_start_date.year}"
                                        inferred_year = statement_start_date.year
                                    else:
                                        date_str = f"{day:02d}/{month:02d}/{statement_end_date.year}"
                                        inferred_year = statement_end_date.year
                    except Exception as e:
                        pass
                
                # Debug: log year determination
                year_debug_entry = {
                    'page': page_num,
                    'line': line_stripped,
                    'raw_date_str': raw_date_str,
                    'current_year_before': current_year,
                    'inferred_year': inferred_year,
                    'date_str_before_chrono': date_str
                }
                
                if inferred_year:
                    # Reject inferred years before 2000 or after current year
                    # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                    if inferred_year < 2000 or inferred_year > CURRENT_YEAR:
                        year_debug_entry['year_validation'] = 'REJECTED'
                        year_debug_entry['reason'] = f'Year {inferred_year} outside valid range (2000-{CURRENT_YEAR})'
                        if current_year is None:
                            current_year = CURRENT_YEAR
                        # Don't update current_year with the invalid year
                        # But still process the date with current_year
                        if date_str:
                            # Re-parse with valid current_year
                            date_str, _ = parse_date_with_year_inference(raw_date_str, current_year, month_map)
                            year_debug_entry['date_str_after_reparse'] = date_str
                    else:
                        year_debug_entry['year_validation'] = 'ACCEPTED'
                        current_year = inferred_year
                    # If this was just a year marker, don't treat as a transaction date
                    # BUT only skip if the line has no amounts (to avoid skipping lines like "2023 25000.00")
                    if date_str is None:
                        year_debug_entry['type'] = 'YEAR_MARKER_ONLY'
                        debug_info['year_determination_debug'].append(year_debug_entry)
                        # Check if this line has amounts before treating as year marker
                        amounts_check = extract_amounts_with_positions(line_words)
                        if len(amounts_check) < 1:
                            all_content.append({
                                'Page': page_num,
                                'Content': f"[YEAR MARKER] {current_year}"
                            })
                            continue
                    else:
                        year_debug_entry['type'] = 'TRANSACTION_DATE'
                        
                        # Use date range to determine year if available
                        if date_range and date_str:
                            try:
                                from datetime import datetime
                                current_date = datetime.strptime(date_str, '%d/%m/%Y')
                                # Check if the date (without year) falls within the date range
                                # Create test dates with both years from the range
                                test_date_start_year = current_date.replace(year=date_range['start_year'])
                                test_date_end_year = current_date.replace(year=date_range['end_year'])
                                
                                # Check which year makes the date fall within the range
                                if date_range['start_date'] <= test_date_start_year <= date_range['end_date']:
                                    # Use start year
                                    date_str = test_date_start_year.strftime('%d/%m/%Y')
                                    current_year = date_range['start_year']
                                    year_debug_entry['date_range_used'] = f'Start year {date_range["start_year"]}'
                                elif date_range['start_date'] <= test_date_end_year <= date_range['end_date']:
                                    # Use end year
                                    date_str = test_date_end_year.strftime('%d/%m/%Y')
                                    current_year = date_range['end_year']
                                    year_debug_entry['date_range_used'] = f'End year {date_range["end_year"]}'
                                else:
                                    # Date doesn't fall within range, use closer year
                                    if abs(current_date.year - date_range['start_year']) < abs(current_date.year - date_range['end_year']):
                                        date_str = test_date_start_year.strftime('%d/%m/%Y')
                                        current_year = date_range['start_year']
                                        year_debug_entry['date_range_used'] = f'Closer to start year {date_range["start_year"]}'
                                    else:
                                        date_str = test_date_end_year.strftime('%d/%m/%Y')
                                        current_year = date_range['end_year']
                                        year_debug_entry['date_range_used'] = f'Closer to end year {date_range["end_year"]}'
                            except:
                                pass  # If date parsing fails, keep as is
                        
                        # Chronological ordering: ensure date is not earlier than last date on page
                        if last_date_on_page and date_str:
                            try:
                                from datetime import datetime
                                current_date = datetime.strptime(date_str, '%d/%m/%Y')
                                last_date = datetime.strptime(last_date_on_page, '%d/%m/%Y')
                                year_debug_entry['last_date_on_page'] = last_date_on_page
                                year_debug_entry['current_date_before_chrono'] = date_str
                                if current_date < last_date:
                                    # Date is earlier, increment year
                                    new_year = current_date.year + 1
                                    if new_year <= CURRENT_YEAR:
                                        date_str = current_date.replace(year=new_year).strftime('%d/%m/%Y')
                                        current_year = new_year
                                        year_debug_entry['chronological_adjustment'] = f'Year incremented from {current_date.year} to {new_year}'
                                        year_debug_entry['date_str_after_chrono'] = date_str
                                    else:
                                        year_debug_entry['chronological_adjustment'] = 'SKIPPED (would exceed CURRENT_YEAR)'
                                else:
                                    year_debug_entry['chronological_adjustment'] = 'NONE (date is chronologically correct)'
                            except:
                                year_debug_entry['chronological_error'] = 'Date parsing failed'
                        else:
                            year_debug_entry['chronological_adjustment'] = 'NONE (no previous date on page)'
                        # Update last date on page
                        last_date_on_page = date_str
                        year_debug_entry['final_date_str'] = date_str
                        year_debug_entry['final_current_year'] = current_year
                        debug_info['year_determination_debug'].append(year_debug_entry)
                
                # Debug: store all lines to see what's happening
                if 'line_processing_debug' not in debug_info:
                    debug_info['line_processing_debug'] = []
                debug_info['line_processing_debug'].append({
                    'page': page_num,
                    'line': line_stripped[:100],
                    'in_transaction_section': in_transaction_section,
                    'date_match': bool(date_match),
                    'raw_date_str': raw_date_str,
                    'date_str': date_str,
                    'inferred_year': inferred_year
                })
                
                # Debug: store date parsing info for debug window (store ALL attempts, even failed ones)
                if 'date_parsing_debug' not in debug_info:
                    debug_info['date_parsing_debug'] = []
                debug_info['date_parsing_debug'].append({
                    'page': page_num,
                    'raw_date_str': raw_date_str,
                    'parsed_date_str': date_str,
                    'line': line_stripped[:100],
                    'date_match': bool(date_match),
                    'date_str_truthy': bool(date_str)
                })
                
                # Find all monetary amounts in this line with their visual x-positions
                amounts = extract_amounts_with_positions(line_words)
                
                # Separate amounts by zone: description vs transaction columns
                desc_zone_amounts = []
                tx_zone_amounts = []
                if column_boundaries and amounts:
                    _, _, desc_left, desc_right, _, _, _, _, _, _ = column_boundaries
                    if desc_left is not None and desc_right is not None:
                        for amt in amounts:
                            x = amt['x_center']
                            in_desc_zone = desc_left <= x <= desc_right
                            if in_desc_zone:
                                desc_zone_amounts.append(amt)
                            else:
                                tx_zone_amounts.append(amt)
                        amounts = tx_zone_amounts  # Only process tx zone amounts for columns
                
                # Fallback: auto-detect columns from first transaction line with multiple amounts
                if column_boundaries is None and len(amounts) >= 2:
                    auto_boundaries = auto_detect_visual_boundaries(amounts)
                    column_boundaries = equalize_debit_credit_ranges(auto_boundaries)
                    debug_info['column_boundaries'] = column_boundaries
                    all_content.append({
                        'Page': page_num,
                        'Content': f"[AUTO-DETECTED COLUMNS] {line_stripped}"
                    })
                
                # Check if this is a description-only line (no amounts)
                if len(amounts) < 1:
                    # Extract the description text
                    desc_only_text = ' '.join(w['text'] for w in line_words)
                    desc_only_text = re.sub(r'^[\s\|\-\.>]+', '', desc_only_text).strip()
                    
                    # Remove date from description if present
                    if date_match and date_str:
                        if raw_date_str and desc_only_text.startswith(raw_date_str):
                            desc_only_text = desc_only_text[len(raw_date_str):].strip()
                        elif raw_date_str and desc_only_text.lower().startswith(raw_date_str.lower()):
                            desc_only_text = desc_only_text[len(raw_date_str):].strip()
                    
                    # Store the date for use in the next transaction (handles multi-line transactions)
                    # Rule: If the new date is before but within 29 days of the previous last_seen_date, don't update
                    # This prevents dates in description lines from overwriting the actual transaction date
                    if date_str:
                        if 'last_seen_date' not in debug_info:
                            debug_info['last_seen_date'] = None
                        
                        # Check if we should update last_seen_date
                        should_update = True
                        if debug_info['last_seen_date']:
                            try:
                                from datetime import datetime
                                prev_date = datetime.strptime(debug_info['last_seen_date'], '%d/%m/%Y')
                                new_date = datetime.strptime(date_str, '%d/%m/%Y')
                                days_diff = (prev_date - new_date).days
                                
                                # If new date is before previous date and within 29 days, don't update
                                if new_date < prev_date and 0 < days_diff <= 29:
                                    should_update = False
                            except:
                                pass  # If date parsing fails, allow update
                        
                        if should_update:
                            debug_info['last_seen_date'] = date_str
                    
                    if not desc_only_text:
                        continue
                    
                    # Check if this is a header line (should NOT be merged with transactions)
                    is_header_only_line = (
                        any(kw in line_lower for kw in ['particulars', 'description', 'details', 'transaction']) and
                        any(kw in line_lower for kw in ['debit', 'credit', 'balance', 'amount'])
                    )
                    if is_header_only_line:
                        all_content.append({
                            'Page': page_num,
                            'Content': f"[COLUMN HEADER] {line_stripped}"
                        })
                        continue
                    
                    # Determine behavior based on multiline_mode
                    force_previous = multiline_mode == "Merge with PREVIOUS transaction"
                    force_next = multiline_mode == "Merge with NEXT transaction"
                    auto_detect = multiline_mode == "Auto-detect"
                    
                    belongs_to_next = force_next  # User forced next
                    
                    if auto_detect and not force_previous:
                        # Look ahead to determine if this belongs to previous or next transaction
                        current_idx = lines.index(line_words)
                        
                        # Check next few lines
                        for lookahead in range(1, 4):
                            if current_idx + lookahead < len(lines):
                                next_line_words = lines[current_idx + lookahead]
                                next_line_text = ' '.join(w['text'] for w in next_line_words).strip()
                                next_amounts = extract_amounts_with_positions(next_line_words)
                                
                                # If next line has amounts, this description belongs to it
                                if len(next_amounts) > 0:
                                    belongs_to_next = True
                                    break
                                # If we hit another description-only line or empty line, stop looking
                                if not next_line_text or len(next_amounts) < 1:
                                    break
                    
                    if belongs_to_next:
                        # Store as pending prefix for next transaction
                        if 'pending_prefix' not in debug_info:
                            debug_info['pending_prefix'] = []
                        debug_info['pending_prefix'].append({
                            'page': page_num,
                            'text': desc_only_text,
                            'will_merge_with_next': True
                        })
                        # Don't add to all_content - it will be merged with next transaction
                        continue
                    else:
                        # This is a continuation of previous transaction
                        if transactions and not date_match:
                            # Append to previous transaction
                            prev_tx = transactions[-1]
                            prev_tx['Transaction Description'] += ' ' + desc_only_text
                            
                            # Also update debug info
                            if debug_info['transactions_debug']:
                                debug_info['transactions_debug'][-1]['description'] += ' ' + desc_only_text
                                debug_info['transactions_debug'][-1]['amounts'].append({
                                    'value': None,
                                    'type': 'continuation_desc',
                                    'text': desc_only_text
                                })
                            continue
                        else:
                            # No previous transaction to append to, treat as regular content
                            all_content.append({
                                'Page': page_num,
                                'Content': line_stripped
                            })
                            continue
                
                # Extract description (excluding amounts assigned to columns unless duplicate exists in desc)
                # Get values of amounts in description zone (to check for duplicates)
                desc_zone_values = {amt['value'] for amt in desc_zone_amounts}
                # Get x-ranges of amounts that will be assigned to columns (tx_zone_amounts)
                tx_amount_ranges = [(amt['x0'], amt['x1'], amt['value']) for amt in tx_zone_amounts] if amounts else []
                
                desc_words = []
                for w in line_words:
                    # Check if this word is an amount assigned to a column
                    is_assigned_amount = False
                    for ax0, ax1, val in tx_amount_ranges:
                        if w['x0'] == ax0 and w['x1'] == ax1:
                            # This is an amount in tx zone - check for duplicate in desc zone
                            if val not in desc_zone_values:
                                # No duplicate in description - exclude it
                                is_assigned_amount = True
                            break
                    # Check if this word is part of the date (exclude from description)
                    is_date_word = False
                    if date_match and raw_date_str:
                        word_text = w['text'].upper()
                        # Check if this word matches the date pattern
                        if re.match(r'^\d{1,2}[A-Z]{3}$', word_text):  # Format like 06NOV
                            is_date_word = True
                        elif re.match(r'^\d{1,2}[/-]\d{1,2}[/-]?\d{0,4}$', word_text):  # Format like 06/11/2023
                            is_date_word = True
                        elif re.match(r'^\d{1,2}\s+[A-Z]{3,9}$', word_text):  # Format like 06 NOV
                            is_date_word = True
                        # Also check if the word is part of the raw_date_str
                        elif raw_date_str and word_text in raw_date_str.upper():
                            is_date_word = True
                    
                    if not is_assigned_amount and not is_date_word:
                        # Filter out "blank" and "None" words
                        if w['text'].lower() not in ['blank', 'none']:
                            desc_words.append(w['text'])
                
                description = ' '.join(desc_words)
                
                # Debug: show what's happening with date exclusion
                if date_match and raw_date_str:
                    if 'date_exclusion_debug' not in debug_info:
                        debug_info['date_exclusion_debug'] = []
                    debug_info['date_exclusion_debug'].append({
                        'page': page_num,
                        'raw_date_str': raw_date_str,
                        'date_str': date_str,
                        'description_after_exclusion': description[:100],
                        'line_words_count': len(line_words),
                        'desc_words_count': len(desc_words)
                    })
                
                # Remove date from description if present
                if date_match and date_str:
                    # Try to remove the raw date string from the beginning of description
                    if raw_date_str and description.startswith(raw_date_str):
                        description = description[len(raw_date_str):].strip()
                    # Also try removing the date if it's at the start (with variations)
                    elif description.lower().startswith(raw_date_str.lower() if raw_date_str else ''):
                        description = description[len(raw_date_str):].strip()
                    # Try removing common date formats from start
                    else:
                        # Remove date pattern from start of description
                        for date_fmt in [r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s*', r'^\d{1,2}\s+[A-Za-z]{3,9}\s*', r'^\d{1,2}[A-Za-z]{3,9}\s*']:
                            new_desc = re.sub(date_fmt, '', description).strip()
                            if new_desc != description:
                                description = new_desc
                                break
                
                description = re.sub(r'^[\s\|\-\.>]+', '', description)
                
                # Check for pending prefix descriptions that should be prepended
                pending_prefixes = []
                if 'pending_prefix' in debug_info and debug_info['pending_prefix']:
                    # Get ALL pending prefixes (from any page) - carry over across page breaks
                    for prefix in debug_info['pending_prefix'][:]:
                        pending_prefixes.append(prefix['text'])
                        debug_info['pending_prefix'].remove(prefix)
                
                if pending_prefixes:
                    # Prepend prefixes to current description
                    full_prefix = ' '.join(pending_prefixes)
                    description = full_prefix + ' ' + description if description else full_prefix
                
                # Assign amounts to columns based on visual x-position
                debit = 0.0
                credit = 0.0
                balance = 0.0
                
                # Debug info for this transaction
                tx_debug = {
                    'page': page_num,
                    'description': description,
                    'prefixes_merged': pending_prefixes,
                    'amounts': [],
                    'assigned': {'debit': 0, 'credit': 0, 'balance': 0}
                }
                
                # Check if this is an opening/closing balance line
                desc_lower = description.lower()
                is_balance_only_desc = any(kw in desc_lower for kw in [
                    'brought forward', 'brought fwd', 'opening balance', 
                    'opening', 'start balance', 'balance bf', 'balance b/f',
                    'carried forward', 'carried fwd', 'closing balance'
                ])
                
                for amt in amounts:
                    x_center = amt['x_center']
                    assigned_col = None
                    debug_checks = []  # Track which checks were performed
                    
                    if column_boundaries:
                        # Unpack all 5 column boundaries: DATE, DESC, DEBIT, CREDIT, BALANCE
                        date_left, date_right, desc_left, desc_right, debit_left, debit_right, credit_left, credit_right, balance_left, balance_right = column_boundaries
                        
                        # FIRST: Check if amount is in balance zone - this takes priority
                        # This prevents balance amounts with CR/DR suffixes from being misclassified
                        if balance_left is not None and balance_right is not None:
                            in_balance_zone = balance_left <= x_center <= balance_right
                            debug_checks.append(f"BALANCE: {balance_left:.1f} <= {x_center:.1f} <= {balance_right:.1f} = {in_balance_zone}")
                            if in_balance_zone:
                                # Use signed_value to preserve negative sign from DB suffix or minus sign
                                balance = amt['signed_value']
                                assigned_col = 'BALANCE (zone)'
                        
                        # SECOND: For amounts NOT in balance zone, check CR/DR suffixes FIRST
                        # This is critical for 4-column format where debit and credit boundaries are the same (amount column)
                        if not assigned_col:
                            type_hint = amt.get('type_hint')
                            if type_hint == 'DEBIT':
                                debit = amt['value']
                                assigned_col = 'DEBIT (DB suffix)'
                                debug_checks.append(f"DEBIT: Assigned by DB suffix, value={amt['value']}")
                            elif type_hint == 'CREDIT':
                                credit = amt['value']
                                assigned_col = 'CREDIT (CR suffix)'
                                debug_checks.append(f"CREDIT: Assigned by CR suffix, value={amt['value']}")
                        
                        # THIRD: Only use x-position zones if no CR/DR suffix and not in balance zone
                        # This is for 5-column format where debit and credit are separate columns
                        # Skip zone checks if debit and credit boundaries are the same (4-column format)
                        if not assigned_col and not (debit_left == credit_left and debit_right == credit_right):
                            if debit_left is not None and debit_right is not None:
                                in_debit_zone = debit_left <= x_center <= debit_right
                                debug_checks.append(f"DEBIT: {debit_left:.1f} <= {x_center:.1f} <= {debit_right:.1f} = {in_debit_zone}")
                                if in_debit_zone:
                                    debit = amt['value']
                                    assigned_col = 'DEBIT (zone)'
                            else:
                                debug_checks.append(f"DEBIT: N/A (no boundaries)")
                            
                            if not assigned_col and credit_left is not None and credit_right is not None:
                                in_credit_zone = credit_left <= x_center <= credit_right
                                debug_checks.append(f"CREDIT: {credit_left:.1f} <= {x_center:.1f} <= {credit_right:.1f} = {in_credit_zone}")
                                if in_credit_zone:
                                    credit = amt['value']
                                    assigned_col = 'CREDIT (zone)'
                            else:
                                debug_checks.append(f"CREDIT: N/A (no boundaries)")
                    
                    if not assigned_col:
                        # Fallback: sort by x-position and assign
                        # First check type hint if available (from DB/CR suffix)
                        type_hint = amt.get('type_hint')
                        if type_hint == 'DEBIT':
                            debit = amt['value']
                            assigned_col = 'DEBIT (DB suffix, fallback)'
                        elif type_hint == 'CREDIT':
                            credit = amt['value']
                            assigned_col = 'CREDIT (CR suffix, fallback)'
                        else:
                            sorted_amts = sorted(amounts, key=lambda x: x['x_center'])
                            amt_idx = sorted_amts.index(amt)
                            
                            is_credit_desc = any(kw in desc_lower for kw in ['deposit', 'credit', 'refund'])
                            
                            if len(sorted_amts) == 1:
                                if is_balance_only_desc:
                                    balance = amt['signed_value']
                                    assigned_col = 'BALANCE (keyword)'
                                elif is_credit_desc:
                                    credit = amt['value']
                                    assigned_col = 'CREDIT (keyword)'
                                else:
                                    debit = amt['value']
                                    assigned_col = 'DEBIT (default)'
                            elif len(sorted_amts) == 2:
                                # Two amounts - could be (debit/credit + balance) OR (debit + credit)
                                if amt_idx == 0:
                                    # First amount - check if debit or credit based on keywords
                                    if is_credit_desc:
                                        credit = amt['value']
                                        assigned_col = 'CREDIT (keyword)'
                                    else:
                                        debit = amt['value']
                                        assigned_col = 'DEBIT (default)'
                                else:
                                    # Second amount - could be credit or balance
                                    # Check if credit keyword in description
                                    if is_credit_desc and debit > 0:
                                        # If we already have a debit, this second amount is likely credit
                                        credit = amt['value']
                                        assigned_col = 'CREDIT (keyword, 2nd)'
                                    elif is_balance_only_desc:
                                        # Definitely balance
                                        balance = amt['signed_value']
                                        assigned_col = 'BALANCE (position)'
                                    else:
                                        # Default to balance, but could be credit
                                        balance = amt['signed_value']
                                        assigned_col = 'BALANCE (position, check if should be CREDIT)'
                            else:
                                if amt_idx == 0:
                                    debit = amt['value']
                                    assigned_col = 'DEBIT (position)'
                                elif amt_idx == 1:
                                    credit = amt['value']
                                    assigned_col = 'CREDIT (position)'
                                elif amt_idx == len(sorted_amts) - 1:
                                    balance = amt['signed_value']
                                    assigned_col = 'BALANCE (position)'
                                else:
                                    if debit == 0:
                                        debit = amt['value']
                                        assigned_col = 'DEBIT (extra)'
                                    else:
                                        assigned_col = 'UNASSIGNED'
                    
                    # Add to debug info
                    amt_debug = {
                        'value': amt['value'],
                        'x_center': round(x_center, 1),
                        'x0': round(amt['x0'], 1),
                        'x1': round(amt['x1'], 1),
                        'assigned_to': assigned_col
                    }
                    if column_boundaries and debug_checks:
                        amt_debug['boundary_checks'] = debug_checks
                    tx_debug['amounts'].append(amt_debug)
                
                tx_debug['assigned'] = {
                    'debit': round(debit, 2) if debit > 0 else None,
                    'credit': round(credit, 2) if credit > 0 else None,
                    'balance': round(balance, 2) if balance != 0 else None
                }
                debug_info['transactions_debug'].append(tx_debug)
                
                # Only add if we have a meaningful transaction
                has_transaction_amount = debit > 0 or credit > 0
                has_balance = balance != 0
                is_valid_transaction = has_transaction_amount or has_balance or (description and len(amounts) > 0) or date_str
                
                if is_valid_transaction:
                    # Use last_seen_date if current line doesn't have a date (handles multi-line transactions)
                    # OR if current date is before but within 29 days of last_seen_date (prevents description dates from overwriting)
                    tx_date = date_str if date_str else debug_info.get('last_seen_date', '')
                    
                    # Check if we should use last_seen_date instead of current date_str
                    if date_str and debug_info.get('last_seen_date'):
                        try:
                            from datetime import datetime
                            current_date = datetime.strptime(date_str, '%d/%m/%Y')
                            last_date = datetime.strptime(debug_info['last_seen_date'], '%d/%m/%Y')
                            days_diff = (last_date - current_date).days
                            
                            # If current date is before last_seen_date and within 29 days, use last_seen_date
                            if current_date < last_date and 0 < days_diff <= 29:
                                tx_date = debug_info['last_seen_date']
                        except:
                            pass  # If date parsing fails, use current date_str
                    
                    tx = {
                        'Page': page_num,
                        'Date': tx_date,
                        'Transaction Description': description,
                        'Debit': round(debit, 2) if debit > 0 else None,
                        'Credit': round(credit, 2) if credit > 0 else None,
                        'Balance': round(balance, 2) if balance != 0 else None
                    }
                    transactions.append(tx)
                    
                    # Clear last_seen_date after using it (for next transaction)
                    if date_str:
                        debug_info['last_seen_date'] = None
                    
                    # Debug: store transaction creation details
                    if 'transaction_creation_debug' not in debug_info:
                        debug_info['transaction_creation_debug'] = []
                    debug_info['transaction_creation_debug'].append({
                        'page': page_num,
                        'date_str': date_str,
                        'date_assigned': tx['Date'],
                        'description': description[:100],
                        'has_debit': debit > 0,
                        'has_credit': credit > 0,
                        'has_balance': balance != 0,
                        'is_valid': is_valid_transaction
                    })
                    
                    # Balance reconciliation check
                    rec = debug_info['reconciliation']
                    tx_num = len(transactions)
                    
                    if balance != 0:
                        if rec['starting_balance'] is None:
                            # First balance - set as starting point
                            rec['starting_balance'] = balance
                            rec['running_balance'] = balance
                            tx_debug['reconciliation'] = {
                                'status': 'STARTING_BALANCE',
                                'expected': balance,
                                'actual': balance,
                                'difference': 0.0
                            }
                        else:
                            # Calculate expected balance
                            expected_balance = rec['running_balance'] - debit + credit
                            difference = round(balance - expected_balance, 2)
                            
                            if abs(difference) > 0.01:  # Allow small rounding errors
                                # Check if this discrepancy is explained by previous cumulative error
                                is_explained = abs(difference - rec['cumulative_error']) < 0.01 and rec['cumulative_error'] != 0
                                
                                # Check if the discrepancy amount appears in the description
                                # This indicates the amount was likely extracted from description text
                                description_amount_found = None
                                abs_difference = abs(difference)
                                # Search for the difference amount in the description
                                amount_patterns = [
                                    rf'\${abs_difference:.2f}',
                                    rf'\${abs_difference:,.2f}',
                                    rf'{abs_difference:.2f}',
                                    rf'{abs_difference:,.2f}'
                                ]
                                for pattern in amount_patterns:
                                    if re.search(pattern, description, re.IGNORECASE):
                                        description_amount_found = abs_difference
                                        break
                                
                                # Update cumulative error
                                rec['cumulative_error'] = round(rec['cumulative_error'] + difference, 2)
                                
                                # Discrepancy found!
                                discrepancy = {
                                    'transaction_num': tx_num,
                                    'page': page_num,
                                    'description': description[:50],
                                    'previous_balance': rec['running_balance'],
                                    'debit': debit,
                                    'credit': credit,
                                    'expected_balance': expected_balance,
                                    'actual_balance': balance,
                                    'difference': difference,
                                    'cumulative_error': rec['cumulative_error'],
                                    'is_explained': is_explained,
                                    'description_amount_found': description_amount_found
                                }
                                rec['discrepancies'].append(discrepancy)
                                
                                if is_explained:
                                    tx_debug['reconciliation'] = {
                                        'status': 'EXPLAINED',
                                        'expected': expected_balance,
                                        'actual': balance,
                                        'difference': difference,
                                        'explanation': f'Matches previous cumulative error of ${rec["cumulative_error"] - difference:.2f}'
                                    }
                                else:
                                    tx_debug['reconciliation'] = {
                                        'status': 'DISCREPANCY',
                                        'expected': expected_balance,
                                        'actual': balance,
                                        'difference': difference,
                                        'cumulative_error': rec['cumulative_error'],
                                        'description_amount_found': description_amount_found,
                                        'warning': f"Amount ${description_amount_found:.2f} found in description - may be false positive" if description_amount_found else None
                                    }
                                    
                                    # AUTO-FIX: If amount found in description, look for duplicate transaction
                                    # and merge them (delete false positive, keep real one with combined description)
                                    # This handles both false debits AND false credits
                                    if description_amount_found and len(transactions) >= 2:
                                        # Look at previous transaction for potential duplicate
                                        prev_tx = transactions[-2]  # Previous transaction
                                        curr_tx = transactions[-1]  # Current (false positive) transaction
                                        
                                        # Check if previous transaction has same date and similar description
                                        same_date = prev_tx['Date'] == date_str
                                        similar_desc = False
                                        
                                        # Clean descriptions for comparison (remove common prefixes/suffixes)
                                        prev_desc_clean = prev_tx['Transaction Description'].lower().replace('payment of ', '').replace('transfer of ', '').replace('deposit of ', '')
                                        curr_desc_clean = description.lower().replace('payment of ', '').replace('transfer of ', '').replace('deposit of ', '')
                                        
                                        # Check if one description contains the other
                                        if prev_desc_clean in curr_desc_clean or curr_desc_clean in prev_desc_clean:
                                            similar_desc = True
                                        # Or check if they share significant words
                                        elif len(set(prev_desc_clean.split()) & set(curr_desc_clean.split())) >= 2:
                                            similar_desc = True
                                        
                                        # Check amounts in both transactions
                                        prev_has_debit = prev_tx['Debit'] is not None and prev_tx['Debit'] > 0
                                        prev_has_credit = prev_tx['Credit'] is not None and prev_tx['Credit'] > 0
                                        prev_has_amount = prev_has_debit or prev_has_credit
                                        
                                        curr_has_debit = debit > 0
                                        curr_has_credit = credit > 0
                                        
                                        # Determine which transaction is the false positive
                                        # Case 1: Previous has no amount, current has amount from description → current is false
                                        # Case 2: Both have amounts, but current's amount matches description → current is false
                                        false_positive_is_current = False
                                        
                                        if same_date and similar_desc:
                                            if not prev_has_amount and (curr_has_debit or curr_has_credit):
                                                # Previous is incomplete, current has the amount from description
                                                false_positive_is_current = True
                                            elif prev_has_amount and (curr_has_debit or curr_has_credit):
                                                # Both have amounts - check if current's amount matches the description amount
                                                # The discrepancy amount tells us which one is wrong
                                                if abs(debit - description_amount_found) < 0.01 or abs(credit - description_amount_found) < 0.01:
                                                    # Current transaction's amount matches what was found in description
                                                    false_positive_is_current = True
                                        
                                        if false_positive_is_current:
                                            # This is a false positive! Merge with previous transaction
                                            # Combine descriptions (take the longer/more detailed one)
                                            combined_desc = prev_tx['Transaction Description']
                                            if len(description) > len(prev_tx['Transaction Description']):
                                                combined_desc = description
                                            
                                            # Determine which values to keep:
                                            # - If previous had no debit but current has debit from description → use current's debit (real one)
                                            # - If previous had no credit but current has credit from description → use current's credit (real one)
                                            # - But if previous already had the amount, keep previous's values
                                            if not prev_has_debit and curr_has_debit:
                                                # False debit extracted from description - don't add it
                                                real_debit = 0
                                                real_credit = credit if curr_has_credit else prev_tx['Credit']
                                            elif not prev_has_credit and curr_has_credit:
                                                # False credit extracted from description - don't add it
                                                real_debit = debit if curr_has_debit else prev_tx['Debit']
                                                real_credit = 0
                                            else:
                                                # Keep previous values
                                                real_debit = prev_tx['Debit']
                                                real_credit = prev_tx['Credit']
                                            
                                            # Update previous transaction
                                            prev_tx['Transaction Description'] = combined_desc
                                            # Only update if we found the real values in current transaction
                                            if real_debit is not None and real_debit > 0:
                                                prev_tx['Debit'] = real_debit
                                            if real_credit is not None and real_credit > 0:
                                                prev_tx['Credit'] = real_credit
                                            prev_tx['Balance'] = balance
                                            
                                            # Remove current false positive transaction
                                            transactions.pop()
                                            
                                            # Update debug info
                                            debug_info['transactions_debug'].pop()
                                            tx_num = len(transactions)  # Reset transaction number
                                            
                                            # Mark as fixed
                                            tx_debug['reconciliation']['auto_fixed'] = True
                                            false_type = "debit" if curr_has_debit else "credit"
                                            tx_debug['reconciliation']['fix_note'] = f"Removed false {false_type} extracted from description, merged with real transaction"
                                            
                                            # Recalculate with corrected values
                                            expected_balance = rec['running_balance'] - prev_tx['Debit'] + prev_tx['Credit']
                                            difference = round(balance - expected_balance, 2)
                                            
                                            # Update discrepancy entry
                                            if abs(difference) <= 0.01:
                                                # Fixed! Update status
                                                discrepancy['status'] = 'AUTO_FIXED'
                                                discrepancy['fix_note'] = f'Removed false {false_type} from description text'
                            else:
                                tx_debug['reconciliation'] = {
                                    'status': 'OK',
                                    'expected': expected_balance,
                                    'actual': balance,
                                    'difference': 0.0,
                                    'cumulative_error': rec['cumulative_error']
                                }
                            
                            # Update running balance to actual extracted value
                            rec['running_balance'] = balance
                    else:
                        # No balance on this transaction - just update running balance for tracking
                        if rec['starting_balance'] is not None:
                            rec['running_balance'] = rec['running_balance'] - debit + credit
                            tx_debug['reconciliation'] = {
                                'status': 'NO_BALANCE',
                                'running_balance': rec['running_balance']
                            }
                        else:
                            tx_debug['reconciliation'] = {'status': 'NO_STARTING_BALANCE'}
                            
                else:
                    all_content.append({
                        'Page': page_num,
                        'Content': line_stripped
                    })
    
    return pd.DataFrame(transactions), pd.DataFrame(all_content)


def parse_transactions_from_ocr(ocr_pages_data, debug_info=None, multiline_mode="Auto-detect"):
    """
    Parse transactions from OCR-extracted page data.
    Similar to parse_transactions_visual but works with OCR output structure.
    
    Args:
        ocr_pages_data: List of dicts from process_pdf_with_ocr, each with 'page_num', 'text', 'words', 'lines'
        debug_info: Optional dict to store debug information
        multiline_mode: How to handle multi-line descriptions
    
    Returns: (transactions_df, all_content_df)
    """
    transactions = []
    all_content = []
    
    if debug_info is None:
        debug_info = {}
    
    debug_info['headers_found'] = []
    debug_info['column_boundaries'] = None
    debug_info['transactions_debug'] = []
    debug_info['reconciliation'] = {
        'starting_balance': None,
        'running_balance': 0.0,
        'discrepancies': [],
        'cumulative_error': 0.0
    }
    
    # Same patterns as parse_transactions_visual
    # Made more strict to avoid matching random numbers in descriptions
    # Only accept 4-digit years starting with 20, or 2-digit years (which will be interpreted as 20xx)
    # Hard-coded month names to prevent random words from being matched
    month_names = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
    month_abbr = r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)'
    month_abbr_mixed = r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
    date_pattern = rf'^((0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](20\d{{2}}|\d{{2}})|\d{{1,2}}\s+{month_names}(?:\s+20\d{{2}})?|\d{{1,2}}\s+{month_abbr_mixed}(?:\s+20\d{{2}})?|20\d{{4}}[/-](0?[1-9]|1[0-2])[/-](0?[1-9]|[12][0-9]|3[01])|\d{{1,2}}\s+{month_names}|\d{{1,2}}{month_abbr})'
    
    date_patterns = [r'\bdate\b', r'\btransaction\s+date\b', r'\bvalue\s+date\b', r'\bposting\s+date\b']
    description_patterns = [r'\bdescription\b', r'\bparticulars\b', r'\bdetails\b', r'\btransaction\b', r'\bnarrative\b', r'\bmemo\b', r'\btransaction\s+details\b']
    debit_patterns = [r'\bdeb(it)?s?\b', r'\bwithdrawals?\b', r'\bpayments?\b', r'\bdr\b']
    credit_patterns = [r'\bcr(edits?)?\b', r'\bdeposits?\b', r'\breceived\b', r'\bcr\b']
    amount_patterns = [r'\bamount\b']
    balance_patterns = [r'\bbal(ance)?\b', r'\brunning\s+bal\b', r'\bavail\w*\s+bal\b']
    
    header_patterns = [
        r'transaction\s+description',
        r'date.*description',
        r'transactions?\s+details?',
        r'activity\s+details?',
    ]
    
    column_boundaries = None
    in_transaction_section = False
    
    # Year tracking for date inference
    current_year = None
    month_map = get_month_map()
    
    # Track last date on current page for chronological ordering
    last_date_on_page = None
    
    for page_data in ocr_pages_data:
        page_num = page_data['page_num']
        
        # Reset last date tracker for new page
        last_date_on_page = None
        lines = page_data['lines']
        
        # Scan for column headers on first page with sufficient content
        if column_boundaries is None and lines:
            lines_to_scan = lines[:10] if page_num > 1 else lines
            for line_idx, line_words in enumerate(lines_to_scan):
                test_boundaries, test_headers = detect_visual_column_boundaries_with_debug(
                    line_words, date_patterns, description_patterns, debit_patterns, credit_patterns, amount_patterns, balance_patterns
                )
                if test_headers and len(test_headers) >= 3:
                    column_boundaries = equalize_debit_credit_ranges(test_boundaries)
                    header_info = test_headers
                    line_text = ' '.join(w['text'] for w in line_words)
                    
                    words_checked = [{'text': w['text'], 'x': round((w['x0']+w['x1'])/2, 1)} for w in line_words]
                    
                    debug_info['headers_found'].append({
                        'page': page_num,
                        'line': line_text,
                        'headers': header_info,
                        'words_checked': words_checked
                    })
                    debug_info['column_boundaries'] = column_boundaries
                    
                    all_content.append({
                        'Page': page_num,
                        'Content': f"[COLUMN HEADER - Line {line_idx+1}] {line_text}"
                    })
                    in_transaction_section = True
                    break
        
        # Process each line
        for line_words in lines:
            line_text = ' '.join(w['text'] for w in line_words)
            line_stripped = line_text.strip()
            if not line_stripped or len(line_stripped) < 3:
                continue
            
            line_lower = line_stripped.lower()
            
            is_section_header = any(re.search(pattern, line_lower) for pattern in header_patterns)
            
            if is_section_header:
                in_transaction_section = True
                all_content.append({
                    'Page': page_num,
                    'Content': f"[HEADER] {line_stripped}"
                })
                continue
            
            if not in_transaction_section:
                # Check for StatementPeriod pattern (e.g., "StatementPeriod 30/12/2023to29/01/2024" or "Statement Period 30/12/2023 to 29/01/2024")
                statement_period_match = re.match(r'Statement\s?Period\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*to\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', line_stripped, re.IGNORECASE)
                if statement_period_match:
                    date1_str = statement_period_match.group(1)
                    date2_str = statement_period_match.group(2)
                    # Try to extract years from both dates
                    year1 = None
                    year2 = None
                    year1_match = re.search(r'(\d{2,4})$', date1_str)
                    year2_match = re.search(r'(\d{2,4})$', date2_str)
                    if year1_match:
                        year1 = int(year1_match.group(1))
                        if year1 < 100:
                            year1 = 2000 + year1
                    if year2_match:
                        year2 = int(year2_match.group(1))
                        if year2 < 100:
                            year2 = 2000 + year2
                    # Use the first date's year as the statement year
                    if year1 and year1 >= 2000 and year1 <= CURRENT_YEAR:
                        current_year = year1
                        all_content.append({
                            'Page': page_num,
                            'Content': f"[STATEMENT PERIOD] {line_stripped} (Year: {current_year})"
                        })
                        continue
                
                # Check for year markers only before section headers (like "2022", "2023")
                # This prevents dates in transactions from being incorrectly flagged as year markers
                year_match = re.match(r'^(20\d{2})\s*$', line_stripped)
                if year_match:
                    detected_year = int(year_match.group(1))
                    # Reject future years (e.g., 2077 from store numbers) and years before 2000
                    # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                    if detected_year > CURRENT_YEAR or detected_year < 2000:
                        if current_year is None:
                            current_year = CURRENT_YEAR
                        # Don't update current_year with the invalid year
                    else:
                        current_year = detected_year
                        all_content.append({
                            'Page': page_num,
                            'Content': f"[YEAR MARKER] {current_year}"
                        })
                    continue
                
                all_content.append({
                    'Page': page_num,
                    'Content': line_stripped
                })
                continue
            
            # Check for date
            raw_date_match = re.match(date_pattern, line_stripped)
            raw_date_str = raw_date_match.group(1) if raw_date_match else None
            
            # Initialize date_str to prevent UnboundLocalError
            date_str = None
            
            # Apply year inference and format to dd/mm/yyyy
            date_str, inferred_year = parse_date_with_year_inference(raw_date_str, current_year, month_map)
            if inferred_year:
                # Reject inferred years before 2000 or after current year
                # Instead of skipping, use previous transaction's year or default to CURRENT_YEAR
                if inferred_year < 2000 or inferred_year > CURRENT_YEAR:
                    if current_year is None:
                        current_year = CURRENT_YEAR
                    # Don't update current_year with the invalid year
                    # But still process the date with current_year
                    if date_str:
                        # Re-parse with valid current_year
                        date_str, _ = parse_date_with_year_inference(raw_date_str, current_year, month_map)
                else:
                    current_year = inferred_year
                if date_str is None:
                    continue
                else:
                    # Chronological ordering: ensure date is not earlier than last date on page
                    if last_date_on_page and date_str:
                        try:
                            from datetime import datetime
                            current_date = datetime.strptime(date_str, '%d/%m/%Y')
                            last_date = datetime.strptime(last_date_on_page, '%d/%m/%Y')
                            if current_date < last_date:
                                # Date is earlier, increment year
                                new_year = current_date.year + 1
                                if new_year <= CURRENT_YEAR:
                                    date_str = current_date.replace(year=new_year).strftime('%d/%m/%Y')
                                    current_year = new_year
                        except:
                            pass  # If date parsing fails, keep as is
                    # Update last date on page
                    last_date_on_page = date_str
            
            # Extract amounts
            amounts = extract_amounts_with_positions(line_words)
            
            # Separate amounts by zone
            desc_zone_amounts = []
            tx_zone_amounts = []
            if column_boundaries and amounts:
                _, _, desc_left, desc_right, _, _, _, _, _, _ = column_boundaries
                if desc_left is not None and desc_right is not None:
                    for amt in amounts:
                        x = amt['x_center']
                        in_desc_zone = desc_left <= x <= desc_right
                        if in_desc_zone:
                            desc_zone_amounts.append(amt)
                        else:
                            tx_zone_amounts.append(amt)
                    amounts = tx_zone_amounts
            
            # Auto-detect columns if needed
            if column_boundaries is None and len(amounts) >= 2:
                auto_boundaries = auto_detect_visual_boundaries(amounts)
                column_boundaries = equalize_debit_credit_ranges(auto_boundaries)
                debug_info['column_boundaries'] = column_boundaries
            
            # Build description
            desc_zone_values = {amt['value'] for amt in desc_zone_amounts}
            tx_amount_ranges = [(amt['x0'], amt['x1'], amt['value']) for amt in tx_zone_amounts] if amounts else []
            
            desc_words = []
            for w in line_words:
                is_assigned_amount = False
                for ax0, ax1, val in tx_amount_ranges:
                    if w['x0'] == ax0 and w['x1'] == ax1:
                        if val not in desc_zone_values:
                            is_assigned_amount = True
                        break
                if not is_assigned_amount:
                    desc_words.append(w['text'])
            
            description = ' '.join(desc_words)
            if raw_date_match:
                description = description[len(raw_date_match.group(0)):].strip()
            description = re.sub(r'^[\s\|\-\.>]+', '', description)
            
            # Assign amounts to columns (same logic as parse_transactions_visual)
            debit = 0.0
            credit = 0.0
            balance = 0.0
            
            # First, check if any amount is in balance zone (if column boundaries exist)
            if column_boundaries and amounts:
                _, _, desc_left, desc_right, debit_left, debit_right, credit_left, credit_right, balance_left, balance_right = column_boundaries
                for amt in amounts:
                    x_center = amt['x_center']
                    # Check if in balance zone first
                    if balance_left is not None and balance_right is not None:
                        in_balance_zone = balance_left <= x_center <= balance_right
                        if in_balance_zone:
                            balance = amt['signed_value']
                            # Remove this amount from further processing
                            amounts = [a for a in amounts if a != amt]
                            break
            
            # Then check for type hint from suffix (DB/CR) for remaining amounts
            if amounts:
                for amt in amounts:
                    type_hint = amt.get('type_hint')
                    if type_hint == 'DEBIT':
                        debit = amt['value']
                    elif type_hint == 'CREDIT':
                        credit = amt['value']
            
            # If no type hints assigned, use positional assignment as fallback
            if debit == 0.0 and credit == 0.0 and amounts:
                sorted_amts = sorted(amounts, key=lambda x: x['x_center'])
                if len(sorted_amts) == 1:
                    balance = sorted_amts[0]['signed_value']
                elif len(sorted_amts) == 2:
                    debit = sorted_amts[0]['value']
                    balance = sorted_amts[1]['signed_value']
                elif len(sorted_amts) >= 3:
                    debit = sorted_amts[0]['value'] if len(sorted_amts) > 0 else 0
                    credit = sorted_amts[1]['value'] if len(sorted_amts) > 1 else 0
                    balance = sorted_amts[-1]['signed_value']
            
            # Build transaction
            is_valid = (debit > 0 or credit > 0 or balance != 0 or 
                       (description and len(description) > 3))
            
            if is_valid and (debit > 0 or credit > 0 or balance != 0):
                transactions.append({
                    'Page': page_num,
                    'Date': date_str if date_str else '',
                    'Transaction Description': description,
                    'Debit': round(debit, 2) if debit > 0 else None,
                    'Credit': round(credit, 2) if credit > 0 else None,
                    'Balance': round(balance, 2) if balance != 0 else None
                })
            else:
                all_content.append({
                    'Page': page_num,
                    'Content': line_stripped
                })
    
    return pd.DataFrame(transactions), pd.DataFrame(all_content)


def group_words_by_line(words, y_tolerance=3):
    """Group words by their y-position into lines."""
    if not words:
        return []
    
    # Sort by y-position (top)
    sorted_words = sorted(words, key=lambda w: (w['top'], w['x0']))
    
    lines = []
    current_line = [sorted_words[0]]
    current_y = sorted_words[0]['top']
    
    for word in sorted_words[1:]:
        if abs(word['top'] - current_y) <= y_tolerance:
            current_line.append(word)
        else:
            # Sort line by x-position and add to lines
            lines.append(sorted(current_line, key=lambda w: w['x0']))
            current_line = [word]
            current_y = word['top']
    
    if current_line:
        lines.append(sorted(current_line, key=lambda w: w['x0']))
    
    return lines


def extract_amounts_with_positions(line_words):
    """Extract monetary amounts from line words with their x-positions.
    Handles formats like: 7.09, $7.09, 7.09DB, 2.30CR, -100.50, 0.99, 0.01
    Also handles CR/DR as separate words after the amount (e.g., "100.00" followed by "CR")
    Excludes numbers starting with 00 (likely IDs/reference numbers) but allows 0.xx for values < 1
    """
    amounts = []
    # Pattern matches: optional currency, optional minus, number, optional DB/CR suffix
    # Allow values < 1 (0.99, 0.01) but exclude pure IDs like 00123
    amount_pattern = r'^[\$€£¥]?\s*(-?)(0?\d[\d,]*\.\d{2})(?:\s*(DB|CR|DR|CRD))?\s*$'
    # Pattern for standalone CR/DR suffix words
    suffix_pattern = r'^(DB|CR|DR|CRD)\s*$'
    
    for i, word in enumerate(line_words):
        text = word['text'].strip()
        match = re.match(amount_pattern, text, re.IGNORECASE)
        if match:
            try:
                has_minus = match.group(1) == '-'
                val = float(match.group(2).replace(',', ''))
                type_hint = match.group(3).upper() if match.group(3) else None
                
                # Check if the next word is a CR/DR suffix (separate word case)
                if not type_hint and i + 1 < len(line_words):
                    next_word = line_words[i + 1]
                    next_text = next_word['text'].strip()
                    suffix_match = re.match(suffix_pattern, next_text, re.IGNORECASE)
                    if suffix_match:
                        type_hint = suffix_match.group(1).upper()
                
                # Normalize type hints: DB/DR -> debit, CR/CRD -> credit
                normalized_hint = None
                if type_hint in ['DB', 'DR']:
                    normalized_hint = 'DEBIT'
                elif type_hint in ['CR', 'CRD']:
                    normalized_hint = 'CREDIT'
                
                # For debit/credit: always use absolute value (column/suffix determines type)
                # For balance: preserve sign from minus or DB suffix
                is_negative = has_minus or normalized_hint == 'DEBIT'
                
                amounts.append({
                    'value': abs(val),  # Absolute value for amount magnitude
                    'signed_value': -val if is_negative else val,  # Signed value for balance
                    'is_negative': is_negative,  # Track if this should be negative
                    'x_center': (word['x0'] + word['x1']) / 2,  # Center x-coordinate
                    'x0': word['x0'],
                    'x1': word['x1'],
                    'text': text,
                    'type_hint': normalized_hint  # DEBIT, CREDIT, or None
                })
            except:
                continue
    
    return amounts


def detect_visual_column_boundaries_with_debug(header_words, date_patterns, description_patterns, debit_patterns, credit_patterns, amount_patterns, balance_patterns):
    """
    Detect ALL column header positions using visual x-coordinates and calculate boundaries.
    Supports both 5-column format (Date, Description, Debit, Credit, Balance) and
    4-column format (Date, Description, Amount, Balance) where Amount has CR/DR suffixes.
    Returns both boundaries (10 values for 5 columns) and debug info about headers found.

    Boundaries format: (date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r)
    For 4-column format: debit and credit boundaries will be the same as amount boundaries.
    """
    date_pos = None
    desc_pos = None
    debit_pos = None
    credit_pos = None
    amount_pos = None
    balance_pos = None
    header_info = []
    
    for word in header_words:
        text_lower = word['text'].lower()
        x_center = (word['x0'] + word['x1']) / 2
        x0, x1 = word['x0'], word['x1']
        
        # Check for date header
        for pattern in date_patterns:
            if re.search(pattern, text_lower):
                date_pos = x_center
                header_info.append({'type': 'DATE', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
        
        # Check for description header
        for pattern in description_patterns:
            if re.search(pattern, text_lower):
                desc_pos = x_center
                header_info.append({'type': 'DESCRIPTION', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
        
        # Check for debit header
        for pattern in debit_patterns:
            if re.search(pattern, text_lower):
                debit_pos = x_center
                header_info.append({'type': 'DEBIT', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
        
        # Check for credit header
        for pattern in credit_patterns:
            if re.search(pattern, text_lower):
                credit_pos = x_center
                header_info.append({'type': 'CREDIT', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
        
        # Check for amount header (for 4-column format with CR/DR suffixes)
        for pattern in amount_patterns:
            if re.search(pattern, text_lower):
                amount_pos = x_center
                header_info.append({'type': 'AMOUNT', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
        
        # Check for balance header
        for pattern in balance_patterns:
            if re.search(pattern, text_lower):
                balance_pos = x_center
                header_info.append({'type': 'BALANCE', 'text': word['text'], 'x_center': round(x_center, 1), 'x0': round(x0, 1), 'x1': round(x1, 1)})
                break
    
    # Calculate ALL 5 column boundaries at midpoints between header centers
    # Order: DATE, DESCRIPTION, DEBIT, CREDIT, BALANCE
    # Returns: (date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r)
    boundaries = None
    
    # Collect all detected positions
    positions = []
    if date_pos: positions.append(('DATE', date_pos))
    if desc_pos: positions.append(('DESC', desc_pos))
    if debit_pos: positions.append(('DEBIT', debit_pos))
    if credit_pos: positions.append(('CREDIT', credit_pos))
    if amount_pos: positions.append(('AMOUNT', amount_pos))
    if balance_pos: positions.append(('BALANCE', balance_pos))
    
    if len(positions) >= 2:
        # Sort by x-position
        positions.sort(key=lambda x: x[1])
        
        # Calculate boundaries at midpoints
        # Default: use 0 for leftmost, extend rightmost to infinity
        # For gaps, estimate proportional spacing
        
        all_bounds = {}
        for i, (col_type, pos) in enumerate(positions):
            if i == 0:
                # Leftmost column
                left = 0
            else:
                # Midpoint between this and previous column
                left = (positions[i-1][1] + pos) / 2
            
            if i == len(positions) - 1:
                # Rightmost column
                right = 10000
            else:
                # Midpoint between this and next column
                right = (pos + positions[i+1][1]) / 2
            
            all_bounds[col_type] = (left, right)
        
        # Build final tuple in order: DATE, DESC, DEBIT, CREDIT, BALANCE
        # Get detected bounds or create estimated ones
        date_bounds = all_bounds.get('DATE')
        desc_bounds = all_bounds.get('DESC')
        debit_bounds = all_bounds.get('DEBIT')
        credit_bounds = all_bounds.get('CREDIT')
        amount_bounds = all_bounds.get('AMOUNT')
        balance_bounds = all_bounds.get('BALANCE')
        
        # If AMOUNT column is detected (4-column format), use it for both debit and credit
        # The CR/DR suffix in the amount will determine debit vs credit
        if amount_bounds and not debit_bounds and not credit_bounds:
            debit_bounds = amount_bounds
            credit_bounds = amount_bounds
        
        # Estimate any missing columns with gaps between zones
        # This ensures no overlap: each zone ends where the next begins
        first_pos = positions[0][1] if positions else 100
        last_pos = positions[-1][1] if positions else 500
        
        if date_bounds is None:
            date_bounds = (0, desc_bounds[0] if desc_bounds else first_pos * 0.5)
        if desc_bounds is None:
            desc_start = date_bounds[1] if date_bounds else 0
            desc_end = debit_bounds[0] if debit_bounds else amount_bounds[0] if amount_bounds else first_pos * 0.8
            desc_bounds = (desc_start, desc_end)
        if debit_bounds is None:
            debit_start = desc_bounds[1] if desc_bounds else date_bounds[1] if date_bounds else 0
            debit_end = credit_bounds[0] if credit_bounds else amount_bounds[0] if amount_bounds else (credit_pos - 0.1) if credit_pos else balance_bounds[0] * 0.7 if balance_bounds else last_pos * 0.6
            debit_bounds = (debit_start, debit_end)
        if credit_bounds is None:
            credit_start = debit_bounds[1] if debit_bounds else debit_pos if debit_pos else amount_bounds[0] if amount_bounds else desc_bounds[1] if desc_bounds else first_pos
            credit_end = balance_bounds[0] if balance_bounds else (balance_pos - 0.1) if balance_pos else last_pos * 0.8
            credit_bounds = (credit_start, credit_end)
        if balance_bounds is None:
            balance_start = credit_bounds[1] if credit_bounds else credit_pos if credit_pos else amount_bounds[1] if amount_bounds else debit_bounds[1] if debit_bounds else last_pos
            balance_bounds = (balance_start, 10000)
        
        boundaries = (date_bounds[0], date_bounds[1], 
                     desc_bounds[0], desc_bounds[1],
                     debit_bounds[0], debit_bounds[1],
                     credit_bounds[0], credit_bounds[1],
                     balance_bounds[0], balance_bounds[1])
    
    elif len(positions) == 1:
        # Only one header found - estimate all other columns
        col_type, pos = positions[0]
        if col_type == 'BALANCE':
            # Estimate: Date (0-20%), Desc (20-40%), Debit (40-55%), Credit (55-70%), Balance (70-100%)
            boundaries = (0, pos*0.2, pos*0.2, pos*0.4, pos*0.4, pos*0.55, pos*0.55, pos*0.7, pos*0.7, 10000)
        elif col_type == 'DATE':
            # Estimate from date position
            boundaries = (0, pos*2, pos*2, pos*4, pos*4, pos*6, pos*6, pos*8, pos*8, 10000)
        else:
            # Generic estimate
            boundaries = (0, 100, 100, 200, 200, 400, 400, 600, 600, 10000)
    
    # Apply equalization to make debit/credit columns have similar width
    boundaries = equalize_debit_credit_ranges(boundaries)
    
    return boundaries, header_info


def detect_visual_column_boundaries(header_words, date_patterns, description_patterns, debit_patterns, credit_patterns, amount_patterns, balance_patterns):
    """
    Detect ALL column header positions using visual x-coordinates and calculate boundaries.
    Supports both 5-column format (Date, Description, Debit, Credit, Balance) and
    4-column format (Date, Description, Amount, Balance) where Amount has CR/DR suffixes.
    Boundaries are at midpoints between adjacent column centers.
    """
    boundaries, _ = detect_visual_column_boundaries_with_debug(
        header_words, date_patterns, description_patterns, debit_patterns, credit_patterns, amount_patterns, balance_patterns
    )
    return boundaries


def equalize_debit_credit_ranges(boundaries):
    """
    Equalize the width of debit and credit columns to prevent description amounts
    from being misclassified. Takes the smaller range and adjusts the wider one.
    
    boundaries: (date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r)
    """
    if boundaries is None:
        return None
    
    date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r = boundaries
    
    # Only adjust if both debit and credit have valid boundaries
    if debit_l is None or debit_r is None or credit_l is None or credit_r is None:
        return boundaries
    
    debit_width = debit_r - debit_l
    credit_width = credit_r - credit_l
    
    # Take the smaller width and apply to both columns
    min_width = min(debit_width, credit_width)
    
    # Adjust debit: keep right boundary, shrink from left if needed
    if debit_width > min_width:
        # Shrink debit column from the left side
        new_debit_l = debit_r - min_width
        debit_l = max(new_debit_l, desc_r if desc_r else 0)  # Don't go into description zone
    
    # Adjust credit: keep left boundary, shrink from right if needed  
    if credit_width > min_width:
        # Shrink credit column from the right side
        new_credit_r = credit_l + min_width
        credit_r = min(new_credit_r, balance_l if balance_l else credit_r)
    
    return (date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r)


def auto_detect_visual_boundaries(amounts):
    """
    Auto-detect column boundaries based on visual x-positions when no header found.
    Always ensures all 5 columns (Date, Description, Debit, Credit, Balance) have boundaries.
    Returns: (date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r)
    """
    if len(amounts) < 1:
        return None
    
    # Sort by x-position
    sorted_amts = sorted(amounts, key=lambda x: x['x_center'])
    positions = [a['x_center'] for a in sorted_amts]
    
    # Calculate raw boundaries based on number of amounts found
    if len(positions) == 1:
        # Only one amount - assume it's balance, estimate all others
        p = positions[0]
        boundaries = (0, p*0.15, p*0.15, p*0.35, p*0.35, p*0.50, p*0.50, p*0.65, p*0.65, 10000)
    
    elif len(positions) == 2:
        # Two amounts - assume: debit/credit + balance
        p0, p1 = positions[0], positions[1]
        date_r = p0 * 0.3
        desc_r = p0 * 0.6
        mid = (p0 + p1) / 2
        credit_right = p1 - 0.1  # Small buffer to separate from balance
        boundaries = (0, date_r, date_r, desc_r, desc_r, mid, credit_right, credit_right, p1, 10000)
    
    elif len(positions) == 3:
        # Three amounts: assume debit, credit, balance
        p0, p1, p2 = positions[0], positions[1], positions[2]
        mid1 = (p0 + p1) / 2
        credit_right = p2 - 0.1  # Ensure credit ends before balance starts
        date_r = p0 * 0.3
        desc_r = p0 * 0.6
        boundaries = (0, date_r, date_r, desc_r, desc_r, mid1, mid1, credit_right, credit_right, p2, 10000)
    
    elif len(positions) >= 4:
        # Four+ amounts - use first 3 as debit/credit/balance estimate
        p0, p1, p2 = positions[0], positions[1], positions[2]
        mid1 = (p0 + p1) / 2
        credit_right = p2 - 0.1
        date_r = p0 * 0.3
        desc_r = p0 * 0.6
        boundaries = (0, date_r, date_r, desc_r, desc_r, mid1, mid1, credit_right, credit_right, p2, 10000)
    
    else:
        return None
    
    # Apply equalization to make debit/credit columns have similar width
    return equalize_debit_credit_ranges(boundaries)

def create_bank_statement_excel_full(transactions_df, all_text_df, tables):
    """
    Create Excel file with multiple sheets:
    - Transactions: Parsed transaction data (5 columns)
    - All Text: Every line of text from the PDF
    - Tables: Each extracted table as a separate sheet
    """
    output = io.BytesIO()
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Sheet 1: Transactions
    ws_trans = workbook.create_sheet(title="Transactions")
    # Include Reconciliation column if present
    base_headers = ['Page', 'Date', 'Transaction Description', 'Debit', 'Credit', 'Balance']
    if 'Reconciliation' in transactions_df.columns:
        headers = base_headers + ['Reconciliation']
    else:
        headers = base_headers
    
    for col, header in enumerate(headers, 1):
        cell = ws_trans.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    if not transactions_df.empty:
        # Define colors for reconciliation status
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light yellow
        orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid") # Light orange
        
        rec_col_idx = None
        if 'Reconciliation' in transactions_df.columns:
            rec_col_idx = list(transactions_df.columns).index('Reconciliation') + 1
        
        date_col_idx = list(transactions_df.columns).index('Date') + 1 if 'Date' in transactions_df.columns else None
        desc_col_idx = list(transactions_df.columns).index('Transaction Description') + 1 if 'Transaction Description' in transactions_df.columns else None
        
        for row_idx, row_data in enumerate(transactions_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_trans.cell(row=row_idx, column=col_idx, value=value)
                
                # Format Transaction Description column as text to preserve leading zeros
                if desc_col_idx and col_idx == desc_col_idx and value:
                    cell.number_format = '@'  # Text format
                
                # Convert date strings to Excel date objects
                if date_col_idx and col_idx == date_col_idx and value:
                    try:
                        # Try to parse date string (handles dd/mm/yyyy, dd-mmm-yyyy, etc.)
                        date_str = str(value).strip()
                        # Try common date formats
                        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y']:
                            try:
                                dt = datetime.strptime(date_str, fmt)
                                # Reject years before 2000
                                if dt.year < 2000:
                                    # Use CURRENT_YEAR instead
                                    dt = dt.replace(year=CURRENT_YEAR)
                                cell.value = dt
                                cell.number_format = 'dd/mm/yyyy'
                                break
                            except ValueError:
                                continue
                    except:
                        pass  # Keep as string if parsing fails
                
                # Try to convert numeric strings to actual numbers
                if value and not (date_col_idx and col_idx == date_col_idx):
                    try:
                        value_str = str(value).strip()
                        # Remove common currency symbols and commas
                        cleaned = value_str.replace('$', '').replace(',', '').replace('€', '').replace('£', '')
                        # Try to convert to float
                        num_value = float(cleaned)
                        cell.value = num_value
                        # Apply number format for amounts (2 decimal places)
                        if '.' in cleaned:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                    except (ValueError, AttributeError):
                        pass  # Keep as string if not a number
                
                # Color code reconciliation column
                if rec_col_idx and col_idx == rec_col_idx and value:
                    value_str = str(value)
                    if '✅' in value_str or 'OK' in value_str:
                        cell.fill = green_fill
                    elif '❌' in value_str or 'Diff' in value_str:
                        cell.fill = red_fill
                    elif '⚠️' in value_str or 'Explained' in value_str:
                        cell.fill = orange_fill
                    elif '🏁' in value_str or 'Starting' in value_str:
                        cell.fill = yellow_fill
    
    # Sheet 2: All Text Lines
    if not all_text_df.empty:
        ws_text = workbook.create_sheet(title="All Text Lines")
        headers = ['Page', 'Content']
        for col, header in enumerate(headers, 1):
            cell = ws_text.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, row_data in enumerate(all_text_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_text.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheets 3+: Each table as a separate sheet
    if tables:
        for i, table_info in enumerate(tables, 1):
            sheet_name = f"Table_{i}_Pg{table_info['page']}"
            # Truncate to Excel's 31 char limit
            sheet_name = sheet_name[:31]
            ws_table = workbook.create_sheet(title=sheet_name)
            
            df = table_info['data']
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws_table.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
            
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws_table.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust all column widths
    for ws in workbook.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    
    return output

def main():
    st.set_page_config(
        page_title="PDF to Excel Converter",
        page_icon="📄",
        layout="centered"
    )
    
    st.title("📄 PDF to Excel Converter")
    st.markdown("Upload a PDF file to extract tables and text into Excel format.")
    
    # Bank statement mode option
    bank_mode = st.checkbox(
        "🏦 Bank Statement Mode",
        help="Enable this for bank statements to get Date, Description, Debit, Credit, Balance columns"
    )
    
    # OCR option for scanned documents
    if OCR_AVAILABLE:
        ocr_mode = st.checkbox(
            "🔍 Enable OCR (for scanned PDFs and images)",
            help="Use Tesseract OCR to extract text from scanned documents or image files (PNG, JPG)"
        )
    else:
        st.info("ℹ️ OCR features disabled. Install Tesseract to enable: `pip install pytesseract pdf2image Pillow`")
        ocr_mode = False
    
    # Multi-line description handling option (only show in bank mode)
    if bank_mode:
        multiline_option = st.radio(
            "📝 Multi-line Descriptions:",
            ["Auto-detect", "Merge with PREVIOUS transaction", "Merge with NEXT transaction"],
            index=0,
            help="How to handle descriptions that span multiple lines"
        )
        
        # Page selection options
        st.subheader("📄 Page Selection")
        page_selection_mode = st.radio(
            "Select pages to process:",
            ["All Pages", "Only Specific Pages", "All Pages - Exclude Specific Pages"],
            index=0,
            help="Choose which pages to include in the extraction"
        )
        
        page_spec = None
        exclude_pages = False
        
        if page_selection_mode == "Only Specific Pages":
            page_spec = st.text_input(
                "Pages to include (e.g., '1-9' or '1,2,3,10,25'):",
                placeholder="1-9",
                help="Enter page numbers as ranges (1-9) or comma-separated list (1,2,3,10,25)"
            )
        elif page_selection_mode == "All Pages - Exclude Specific Pages":
            page_spec = st.text_input(
                "Pages to exclude (e.g., '1-9' or '1,2,3,10,25'):",
                placeholder="1",
                help="Enter page numbers to skip as ranges (1-9) or comma-separated list (1,2,3,10,25)"
            )
            exclude_pages = True
    else:
        multiline_option = "Auto-detect"
        page_spec = None
        exclude_pages = False
    
    # Determine accepted file types based on OCR mode
    if ocr_mode:
        accepted_types = ["pdf", "png", "jpg", "jpeg"]
        upload_help = "Upload a PDF or image file (PNG, JPG) to convert"
    else:
        accepted_types = ["pdf"]
        upload_help = "Upload a PDF file to convert"
    
    uploaded_file = st.file_uploader(
        "Choose a file",
        type=accepted_types,
        help=upload_help
    )
    
    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        # Convert button - only process when clicked
        convert_clicked = st.button("🔄 Convert PDF", type="primary", use_container_width=True)
        
        if convert_clicked:
            with st.spinner("Extracting data from PDF..."):
                try:
                    # Only use bank statement mode if explicitly enabled
                    is_bank = bank_mode
                    
                    if is_bank:
                        st.info("🏦 Bank statement mode active. Extracting ALL content with transaction parsing.")
                    
                    # Check if it's an image file
                    is_image = uploaded_file.name.lower().endswith(('.png', '.jpg', '.jpeg'))
                    
                    if is_image and ocr_mode:
                        # Process image directly with OCR
                        st.info("🔍 Processing image with OCR...")
                        image = Image.open(uploaded_file)
                        ocr_text = ocr_image_to_text(image)
                        
                        # Create simple DataFrame from OCR text
                        lines = ocr_text.split('\n')
                        all_text_df = pd.DataFrame([{'Page': 1, 'Content': line} for line in lines if line.strip()])
                        tables = []
                        
                        # Parse transactions from OCR text (simplified)
                        debug_info = {}
                        transactions_df = pd.DataFrame()  # Simplified - would need full OCR parser
                        st.warning("⚠️ Image OCR returns raw text. Full transaction parsing from images is limited in this version.")
                        
                    elif ocr_mode:
                        # OCR mode for PDF - check if scanned and process
                        uploaded_file.seek(0)
                        is_scanned = is_scanned_pdf(uploaded_file)
                        
                        if is_scanned:
                            st.info("🔍 Scanned PDF detected. Running OCR...")
                            uploaded_file.seek(0)
                            ocr_pages = process_pdf_with_ocr(uploaded_file, page_filter)
                            
                            # Build all_text_df from OCR results
                            all_text_lines = []
                            for page_data in ocr_pages:
                                for line in page_data['text'].split('\n'):
                                    if line.strip():
                                        all_text_lines.append({'Page': page_data['page_num'], 'Content': line.strip()})
                            all_text_df = pd.DataFrame(all_text_lines)
                            tables = []  # OCR doesn't extract tables automatically
                            
                            # Parse transactions from OCR data
                            debug_info = {}
                            transactions_df, _ = parse_transactions_from_ocr(ocr_pages, debug_info, multiline_option)
                            
                            if not transactions_df.empty:
                                st.success(f"✅ OCR complete! Extracted {len(transactions_df)} transactions")
                        else:
                            # Not scanned - use regular processing
                            st.info("ℹ️ PDF contains selectable text. Using standard extraction (OCR not needed).")
                            
                            # Extract tables and text normally
                            uploaded_file.seek(0)
                            tables = extract_tables_from_pdf(uploaded_file)
                            
                            uploaded_file.seek(0)
                            all_text_df = extract_text_to_dataframe(uploaded_file)
                            
                            # Parse transactions normally
                            uploaded_file.seek(0)
                            debug_info = {}
                            transactions_df, _ = parse_transactions_visual(uploaded_file, debug_info, multiline_option, page_filter)
                    else:
                        # No OCR - standard processing
                        # Extract ALL tables (like regular mode)
                        uploaded_file.seek(0)
                        tables = extract_tables_from_pdf(uploaded_file)
                        
                        # Extract ALL text (like regular mode)
                        uploaded_file.seek(0)
                        all_text_df = extract_text_to_dataframe(uploaded_file)
                        
                        # Build page filter based on user input
                        page_filter = None
                        if page_spec and page_spec.strip():
                            # First get total pages
                            uploaded_file.seek(0)
                            with pdfplumber.open(uploaded_file) as pdf:
                                total_pages = len(pdf.pages)
                            
                            # Parse the page specification
                            specified_pages = parse_page_spec(page_spec, total_pages)
                            
                            if exclude_pages:
                                # Exclude mode: all pages EXCEPT specified ones
                                page_filter = set(range(1, total_pages + 1)) - specified_pages
                                st.info(f"📄 Excluding pages: {sorted(specified_pages)} | Processing pages: {sorted(page_filter)}")
                            else:
                                # Include mode: only specified pages
                                page_filter = specified_pages
                                st.info(f"📄 Processing pages: {sorted(page_filter)}")
                        
                        # Parse transactions using VISUAL coordinate detection with debug info
                        uploaded_file.seek(0)
                        debug_info = {}
                        transactions_df, _ = parse_transactions_visual(uploaded_file, debug_info, multiline_option, page_filter)
                    
                    # Show tabs for different views
                    tab1, tab2, tab3, tab4 = st.tabs(["📊 Transactions", "📄 All Text Lines", "📋 Tables", "🔍 Debug View"])
                    
                    with tab1:
                        if not transactions_df.empty:
                            st.write(f"**Found {len(transactions_df)} transaction(s)**")
                            
                            # Show reconciliation warning if discrepancies found
                            if debug_info.get('reconciliation') and debug_info['reconciliation']['discrepancies']:
                                num_disc = len(debug_info['reconciliation']['discrepancies'])
                                st.error(f"⚠️ Reconciliation Alert: {num_disc} balance discrepancy(ies) detected! Check 🔍 Debug View for details.")
                                
                                # Check if any discrepancies have description amounts
                                desc_amount_count = sum(1 for d in debug_info['reconciliation']['discrepancies'] if d.get('description_amount_found'))
                                if desc_amount_count > 0:
                                    st.warning(f"📝 Note: {desc_amount_count} discrepancy(s) have amounts found in transaction descriptions - these may be false positives extracted from text.")
                            elif debug_info.get('reconciliation') and debug_info['reconciliation']['starting_balance'] is not None:
                                st.success("✅ Balances reconcile correctly")
                            
                            # Add reconciliation column to DataFrame for Excel export
                            if debug_info.get('reconciliation'):
                                rec_statuses = []
                                for i, tx in enumerate(debug_info['transactions_debug'], 1):
                                    rec = tx.get('reconciliation', {})
                                    status = rec.get('status', '')
                                    if status == 'STARTING_BALANCE':
                                        rec_statuses.append('🏁 Starting')
                                    elif status == 'OK':
                                        rec_statuses.append('✅ OK')
                                    elif status == 'DISCREPANCY':
                                        diff = rec.get('difference', 0)
                                        rec_statuses.append(f'❌ Diff: ${diff:.2f}')
                                    elif status == 'EXPLAINED':
                                        diff = rec.get('difference', 0)
                                        rec_statuses.append(f'⚠️ Explained: ${diff:.2f}')
                                    elif status == 'NO_BALANCE':
                                        rec_statuses.append('ℹ️ No Balance')
                                    else:
                                        rec_statuses.append('')
                                # Pad if needed
                                while len(rec_statuses) < len(transactions_df):
                                    rec_statuses.append('')
                                transactions_df = transactions_df.copy()
                                transactions_df['Reconciliation'] = rec_statuses[:len(transactions_df)]
                            
                            # Show summary stats
                            total_debit = pd.to_numeric(transactions_df['Debit'], errors='coerce').sum()
                            total_credit = pd.to_numeric(transactions_df['Credit'], errors='coerce').sum()
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Total Debit", f"${total_debit:,.2f}")
                            col2.metric("Total Credit", f"${total_credit:,.2f}")
                            col3.metric("Net", f"${total_credit - total_debit:,.2f}")
                            
                            st.dataframe(transactions_df, width='stretch')
                        else:
                            st.warning("No transactions detected. The PDF may not contain debit/credit values in expected format.")
                            st.info("💡 Check the 'All Text Lines' tab to see what was extracted from the PDF.")
                    
                    with tab2:
                        if not all_text_df.empty:
                            st.write(f"**All text lines ({len(all_text_df)} items)**")
                            st.dataframe(all_text_df, width='stretch')
                        else:
                            st.info("No text content found.")
                    
                    with tab3:
                        if tables:
                            st.write(f"**Found {len(tables)} table(s)**")
                            for i, table_info in enumerate(tables[:5], 1):  # Show first 5 tables
                                with st.expander(f"Table {i} (Page {table_info['page']}) - Preview"):
                                    st.dataframe(table_info['data'].head(10))
                        else:
                            st.info("No tables detected in the PDF.")
                    
                    with tab4:
                        st.subheader("🔍 Column Detection Debug View")
                        st.write(f"**Multi-line Mode:** {multiline_option}")
                        
                        # Show detected headers
                        if debug_info.get('headers_found'):
                            st.write("**📍 Headers Detected:**")
                            for header in debug_info['headers_found']:
                                st.write(f"Page {header['page']}: `{header['line']}`")
                                
                                # Show all words checked for header detection
                                if header.get('words_checked'):
                                    with st.expander(f"🔍 Words checked on this line ({len(header['words_checked'])} words)"):
                                        words_df = pd.DataFrame(header['words_checked'])
                                        st.dataframe(words_df, width='stretch')
                                        st.write("**Looking for patterns:**")
                                        st.code("""Debit: \\bdeb(it)?\\b, \\bwithdrawal\\b, \\bpayments?\\b
Credit: \\bcr(edits?)?\\b, \\bdeposits?\\b, \\breceived\\b
Balance: \\bbal(ance)?\\b, \\brunning\\s+bal""")
                                
                                if header['headers']:
                                    # Show headers with positions
                                    st.write("✅ **Headers Found:**")
                                    header_df = pd.DataFrame(header['headers'])
                                    st.dataframe(header_df, width='stretch')
                                else:
                                    st.warning("❌ No headers found - check 'Words checked' above to see what text was on this line")
                                    
                                    # Visual diagram of header positions
                                    st.write("**Header Layout on PDF:**")
                                    header_line = ""
                                    positions_line = ""
                                    for h in sorted(header['headers'], key=lambda x: x['x_center']):
                                        label = h['type']
                                        pos = f"{h['x_center']:.0f}"
                                        # Create spacing based on x position
                                        spacing = int(h['x_center'] / 10)
                                        header_line += " " * (spacing - len(header_line)) + f"[{label}]"
                                        positions_line += " " * (spacing - len(positions_line)) + f"  {pos}"
                                    st.code(f"PDF Position (x-coordinates):\n{header_line}\n{positions_line}")
                        else:
                            st.warning("No headers detected - using auto-detection")
                        
                        # Show column boundaries
                        boundaries = debug_info.get('column_boundaries')
                        if boundaries:
                            st.write("**📐 Imaginary Column Boundaries (x-coordinates):**")
                            # Unpack all 5 columns: DATE, DESC, DEBIT, CREDIT, BALANCE
                            date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r = boundaries
                            
                            # Check if any boundaries are None (indicates missing headers)
                            has_debit = debit_l is not None and debit_r is not None
                            has_credit = credit_l is not None and credit_r is not None
                            has_balance = balance_l is not None and balance_r is not None
                            has_date = date_l is not None and date_r is not None
                            has_desc = desc_l is not None and desc_r is not None
                            
                            st.write("**Column Detection Status:**")
                            cols = st.columns(5)
                            with cols[0]:
                                if has_date:
                                    st.success("✅ DATE")
                                else:
                                    st.warning("⚠️ DATE (est)")
                            with cols[1]:
                                if has_desc:
                                    st.success("✅ DESC")
                                else:
                                    st.warning("⚠️ DESC (est)")
                            with cols[2]:
                                if has_debit:
                                    st.success("✅ DEBIT")
                                else:
                                    st.warning("⚠️ DEBIT (est)")
                            with cols[3]:
                                if has_credit:
                                    st.success("✅ CREDIT")
                                else:
                                    st.warning("⚠️ CREDIT (est)")
                            with cols[4]:
                                if has_balance:
                                    st.success("✅ BALANCE")
                                else:
                                    st.warning("⚠️ BALANCE (est)")
                            
                            # Show calculation from headers
                            if debug_info.get('headers_found'):
                                all_detected_headers = []
                                for header in debug_info['headers_found']:
                                    if header.get('headers'):
                                        all_detected_headers.extend(header['headers'])
                                
                                if all_detected_headers:
                                    st.write("**📍 Detected Headers in PDF:**")
                                    sorted_headers = sorted(all_detected_headers, key=lambda x: x['x_center'])
                                    for i, h in enumerate(sorted_headers):
                                        st.write(f"  • **{h['type']}** at x={h['x_center']:.1f} (text: '{h['text']}')")
                                    
                                    if len(sorted_headers) >= 2:
                                        st.write("**Boundary Calculation:**")
                                        st.write("Midpoints between adjacent headers become zone boundaries:")
                                        for i in range(len(sorted_headers) - 1):
                                            mid = (sorted_headers[i]['x_center'] + sorted_headers[i+1]['x_center']) / 2
                                            st.write(f"  → {sorted_headers[i]['type']} zone ends at x={mid:.1f}, {sorted_headers[i+1]['type']} zone starts at x={mid:.1f}")
                            
                            # Create visual representation for all 5 columns
                            col_data = []
                            if date_l is not None and date_r is not None:
                                col_data.append({
                                    'Column': 'DATE' + (' (est.)' if not has_date else ''), 
                                    'Left': round(date_l, 1), 
                                    'Right': round(date_r, 1)
                                })
                            if desc_l is not None and desc_r is not None:
                                col_data.append({
                                    'Column': 'DESCRIPTION' + (' (est.)' if not has_desc else ''), 
                                    'Left': round(desc_l, 1), 
                                    'Right': round(desc_r, 1)
                                })
                            if debit_l is not None and debit_r is not None:
                                col_data.append({
                                    'Column': 'DEBIT' + (' (est.)' if not has_debit else ''), 
                                    'Left': round(debit_l, 1), 
                                    'Right': round(debit_r, 1)
                                })
                            if credit_l is not None and credit_r is not None:
                                col_data.append({
                                    'Column': 'CREDIT' + (' (est.)' if not has_credit else ''), 
                                    'Left': round(credit_l, 1), 
                                    'Right': round(credit_r, 1)
                                })
                            if balance_l is not None and balance_r is not None:
                                col_data.append({
                                    'Column': 'BALANCE' + (' (est.)' if not has_balance else ''), 
                                    'Left': round(balance_l, 1), 
                                    'Right': round(balance_r, 1)
                                })
                            
                            if col_data:
                                st.dataframe(pd.DataFrame(col_data), width='stretch')
                                if not (has_date and has_desc and has_debit and has_credit and has_balance):
                                    st.info("ℹ️ Some boundaries are estimated based on detected headers/amounts. All 5 columns will still have defined ranges.")
                            
                            # Show how assignment works
                            st.write("**🎯 How Amounts Are Assigned:**")
                            st.write(f"""
Each word's **center x-position** is compared to the boundaries above:
- If `date_left ≤ x_center ≤ date_right` → **DATE** (detected: {has_date})
- If `desc_left ≤ x_center ≤ desc_right` → **DESCRIPTION** (detected: {has_desc})  
- If `debit_left ≤ x_center ≤ debit_right` → **DEBIT** (detected: {has_debit})
- If `credit_left ≤ x_center ≤ credit_right` → **CREDIT** (detected: {has_credit})
- If `balance_left ≤ x_center ≤ balance_right` → **BALANCE** (detected: {has_balance})

*Note: (detected: True) means header was found in PDF. (detected: False) means boundary was estimated.*
""")
                            
                            # ASCII art visualization for all 5 columns
                            st.code(f"""
    PDF Page Layout (x-coordinates):
    0 {'─' * 70} 700+ points
    
    {'[DATE]' if date_l else '[    ]':^12} {'[DESC]' if desc_l else '[    ]':^12} {'[DEBIT]' if debit_l else '[     ]':^12} {'[CREDIT]' if credit_l else '[      ]':^12} {'[BALANCE]':^12}
    {'├' + '─' * 10 + '┤' if date_l else ' ' * 12:^12} {'├' + '─' * 10 + '┤' if desc_l else ' ' * 12:^12} {'├' + '─' * 10 + '┤' if debit_l else ' ' * 12:^12} {'├' + '─' * 10 + '┤' if credit_l else ' ' * 12:^12} {'├' + '─' * 10 + '┤':^12}
    {f'{date_l:.0f}-{date_r:.0f}' if date_l else '':^12} {f'{desc_l:.0f}-{desc_r:.0f}' if desc_l else '':^12} {f'{debit_l:.0f}-{debit_r:.0f}' if debit_l else '':^12} {f'{credit_l:.0f}-{credit_r:.0f}' if credit_l else '':^12} {f'{balance_l:.0f}-{balance_r:.0f}':^12}
                            """)
                        
                        # Show date parsing debug details
                        if debug_info.get('date_parsing_debug'):
                            # Count successful vs failed parses
                            successful = sum(1 for d in debug_info['date_parsing_debug'] if d.get('date_str_truthy'))
                            failed = len(debug_info['date_parsing_debug']) - successful
                            st.write(f"**📅 Date Parsing Details ({len(debug_info['date_parsing_debug'])} total, {successful} successful, {failed} failed):**")
                            with st.expander("View date parsing details (first 50)"):
                                for i, date_info in enumerate(debug_info['date_parsing_debug'][:50], 1):
                                    status = "✅" if date_info.get('date_str_truthy') else "❌"
                                    st.write(f"**Date {i} (Page {date_info['page']}) {status}:**")
                                    st.write(f"  Date match found: {date_info.get('date_match')}")
                                    st.write(f"  Raw: `{date_info['raw_date_str']}` → Parsed: `{date_info['parsed_date_str']}`")
                                    st.write(f"  Line: `{date_info['line']}`")
                                    st.write("---")
                        
                        # Show date exclusion debug details
                        if debug_info.get('date_exclusion_debug'):
                            st.write(f"**🔍 Date Exclusion from Description Details ({len(debug_info['date_exclusion_debug'])} dates):**")
                            with st.expander("View date exclusion details (first 50)"):
                                for i, excl_info in enumerate(debug_info['date_exclusion_debug'][:50], 1):
                                    st.write(f"**Date {i} (Page {excl_info['page']}):**")
                                    st.write(f"  Raw: `{excl_info['raw_date_str']}` → Parsed: `{excl_info['date_str']}`")
                                    st.write(f"  Line words: {excl_info['line_words_count']} → Desc words: {excl_info['desc_words_count']}")
                                    st.write(f"  Description after exclusion: `{excl_info['description_after_exclusion']}`")
                                    st.write("---")
                        
                        # Show transaction creation debug details
                        if debug_info.get('transaction_creation_debug'):
                            st.write(f"**🔧 Transaction Creation Details ({len(debug_info['transaction_creation_debug'])} transactions):**")
                            with st.expander("View transaction creation details (first 50)"):
                                for i, tx_info in enumerate(debug_info['transaction_creation_debug'][:50], 1):
                                    st.write(f"**Transaction {i} (Page {tx_info['page']}):**")
                                    st.write(f"  date_str: `{tx_info['date_str']}`")
                                    st.write(f"  date_assigned: `{tx_info['date_assigned']}`")
                                    st.write(f"  description: `{tx_info['description']}`")
                                    st.write(f"  has_debit: {tx_info['has_debit']}, has_credit: {tx_info['has_credit']}, has_balance: {tx_info['has_balance']}")
                                    st.write(f"  is_valid: {tx_info['is_valid']}")
                                    st.write("---")
                        
                        # Show line processing debug details
                        if debug_info.get('line_processing_debug'):
                            st.write(f"**📝 Line Processing Details ({len(debug_info['line_processing_debug'])} lines):**")
                            with st.expander("View line processing details (first 50 lines)"):
                                for i, line_info in enumerate(debug_info['line_processing_debug'][:50], 1):
                                    status = "✅" if line_info['in_transaction_section'] else "❌"
                                    date_status = "✅" if line_info['date_match'] else "❌"
                                    st.write(f"**Line {i} (Page {line_info['page']}) {status} {date_status}:**")
                                    st.write(f"  Line: `{line_info['line']}`")
                                    st.write(f"  in_transaction_section: {line_info['in_transaction_section']}")
                                    st.write(f"  date_match: {line_info['date_match']}")
                                    st.write(f"  raw_date_str: `{line_info['raw_date_str']}`")
                                    st.write("---")
                        
                        # Show transaction debug details
                        if debug_info.get('transactions_debug'):
                            st.write(f"**📊 Transaction Categorization Details ({len(debug_info['transactions_debug'])} transactions):**")
                            
                            # Count merged descriptions
                            prefix_count = sum(1 for tx in debug_info['transactions_debug'] if tx.get('prefixes_merged'))
                            continuation_count = sum(1 for tx in debug_info['transactions_debug'] 
                                             if any(a.get('type') == 'continuation_desc' for a in tx['amounts']))
                            if prefix_count > 0 or continuation_count > 0:
                                st.info(f"ℹ️ Multi-line descriptions: {prefix_count} prefix(es) merged, {continuation_count} continuation(s) merged")
                            
                            # Show first 10 transactions with detailed breakdown
                            for i, tx in enumerate(debug_info['transactions_debug'][:10], 1):
                                # Check if this has merged descriptions
                                has_continuation = any(a.get('type') == 'continuation_desc' for a in tx['amounts'])
                                has_prefix = bool(tx.get('prefixes_merged'))
                                
                                # Add reconciliation indicator
                                rec_status = tx.get('reconciliation', {}).get('status', '')
                                rec_indicator = ''
                                if rec_status == 'DISCREPANCY':
                                    diff = tx['reconciliation'].get('difference', 0)
                                    rec_indicator = f" ❌(${diff:.2f})"
                                elif rec_status == 'EXPLAINED':
                                    diff = tx['reconciliation'].get('difference', 0)
                                    rec_indicator = f" ⚠️(${diff:.2f})"
                                elif rec_status == 'OK':
                                    rec_indicator = " ✅"
                                elif rec_status == 'STARTING_BALANCE':
                                    rec_indicator = " 🏁(start)"
                                
                                expander_label = f"Transaction {i}: {tx['description'][:50]}..."
                                if has_prefix or has_continuation:
                                    expander_label += " 📝"
                                expander_label += rec_indicator
                                
                                with st.expander(expander_label):
                                    st.write(f"**Description:** {tx['description']}")
                                    st.write(f"**Page:** {tx['page']}")
                                    
                                    # Show active column boundaries used for this transaction
                                    if debug_info.get('column_boundaries'):
                                        boundaries = debug_info['column_boundaries']
                                        # Unpack all 5 columns: DATE, DESC, DEBIT, CREDIT, BALANCE
                                        date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r = boundaries
                                        st.write("**📐 Active Column Boundaries:**")
                                        bounds_data = []
                                        if date_l is not None and date_r is not None:
                                            bounds_data.append({'Column': 'DATE', 'Left': f"{date_l:.1f}", 'Right': f"{date_r:.1f}"})
                                        if desc_l is not None and desc_r is not None:
                                            bounds_data.append({'Column': 'DESC', 'Left': f"{desc_l:.1f}", 'Right': f"{desc_r:.1f}"})
                                        if debit_l is not None and debit_r is not None:
                                            bounds_data.append({'Column': 'DEBIT', 'Left': f"{debit_l:.1f}", 'Right': f"{debit_r:.1f}"})
                                        if credit_l is not None and credit_r is not None:
                                            bounds_data.append({'Column': 'CREDIT', 'Left': f"{credit_l:.1f}", 'Right': f"{credit_r:.1f}"})
                                        if balance_l is not None and balance_r is not None:
                                            bounds_data.append({'Column': 'BALANCE', 'Left': f"{balance_l:.1f}", 'Right': f"{balance_r:.1f}"})
                                        if bounds_data:
                                            st.dataframe(pd.DataFrame(bounds_data), width='stretch')
                                    
                                    # Show prefix lines (prepended)
                                    if tx.get('prefixes_merged'):
                                        st.write("**⬆️ Prefix Lines Merged:**")
                                        for line in tx['prefixes_merged']:
                                            st.write(f"  → {line}")
                                    
                                    # Amounts detected
                                    if tx['amounts']:
                                        st.write("**💰 Amounts Detected:**")
                                        # Filter out continuation markers for display
                                        display_amounts = [a for a in tx['amounts'] if a.get('value') is not None]
                                        if display_amounts:
                                            amt_df = pd.DataFrame(display_amounts)
                                            st.dataframe(amt_df, width='stretch')
                                            
                                            # Show position check against boundaries
                                            if debug_info.get('column_boundaries'):
                                                boundaries = debug_info['column_boundaries']
                                                # Unpack all 5 columns
                                                date_l, date_r, desc_l, desc_r, debit_l, debit_r, credit_l, credit_r, balance_l, balance_r = boundaries
                                                st.write("**📍 Position Check:**")
                                                for amt in display_amounts:
                                                    x = amt['x_center']
                                                    assigned = amt.get('assigned_to', 'UNASSIGNED')
                                                    # Show boundary checks if available
                                                    if 'boundary_checks' in amt:
                                                        st.write(f"  ${amt['value']:,.2f} at x={x:.1f} → **{assigned}**")
                                                        for check in amt['boundary_checks']:
                                                            st.write(f"    - {check}")
                                                    else:
                                                        # Fallback: determine which zone it's in (only for amount columns)
                                                        if debit_l is not None and debit_r is not None and debit_l <= x <= debit_r:
                                                            zone = f"DEBIT zone ({debit_l:.0f}-{debit_r:.0f})"
                                                        elif credit_l is not None and credit_r is not None and credit_l <= x <= credit_r:
                                                            zone = f"CREDIT zone ({credit_l:.0f}-{credit_r:.0f})"
                                                        elif balance_l is not None and balance_r is not None and balance_l <= x <= balance_r:
                                                            zone = f"BALANCE zone ({balance_l:.0f}-{balance_r:.0f})"
                                                        else:
                                                            zone = "OUTSIDE zones"
                                                        st.write(f"  ${amt['value']:,.2f} at x={x:.1f} → {zone} → **{assigned}**")
                                        
                                        # Show continuation lines (appended)
                                        continuation_lines = [a for a in tx['amounts'] if a.get('type') == 'continuation_desc']
                                        if continuation_lines:
                                            st.write("**⬇️ Continuation Lines Merged:**")
                                            for line in continuation_lines:
                                                st.write(f"  → {line.get('text', '')}")
                                    
                                    # Reconciliation info
                                    if tx.get('reconciliation'):
                                        st.write("**💰 Reconciliation:**")
                                        rec = tx['reconciliation']
                                        if rec['status'] == 'DISCREPANCY':
                                            st.error(f"❌ Discrepancy: Expected ${rec['expected']:,.2f}, Got ${rec['actual']:,.2f} (Diff: ${rec['difference']:,.2f})")
                                            if 'cumulative_error' in rec:
                                                st.write(f"   Cumulative error: ${rec['cumulative_error']:,.2f}")
                                            # Show warning if amount was found in description
                                            if rec.get('description_amount_found'):
                                                st.warning(f"⚠️ Amount ${rec['description_amount_found']:.2f} found in transaction description! This may be a false positive debit/credit extracted from text.")
                                        elif rec['status'] == 'EXPLAINED':
                                            st.warning(f"⚠️ Explained: Expected ${rec['expected']:,.2f}, Got ${rec['actual']:,.2f}")
                                            st.write(f"   {rec.get('explanation', '')}")
                                        elif rec['status'] == 'OK':
                                            st.success(f"✅ Balances match: ${rec['actual']:,.2f}")
                                        elif rec['status'] == 'STARTING_BALANCE':
                                            st.info(f"🏁 Starting balance: ${rec['actual']:,.2f}")
                                        elif rec['status'] == 'NO_BALANCE':
                                            st.write(f"ℹ️ Running balance: ${rec['running_balance']:,.2f}")
                                    
                                    # Final assignment
                                    st.write("**✅ Final Assignment:**")
                                    st.json(tx['assigned'])
                        else:
                            st.info("No transaction debug data available.")
                        
                        # Show year determination debug
                        if debug_info.get('year_determination_debug'):
                            st.write("---")
                            st.subheader("📅 Year Determination Debug")
                            
                            # Show date range information first
                            date_ranges = [d for d in debug_info['year_determination_debug'] if d['type'] == 'DATE_RANGE_DETECTED']
                            if date_ranges:
                                st.write("**📊 Date Ranges Detected:**")
                                for dr in date_ranges:
                                    st.info(f"Page {dr['page']}: '{dr['line']}'")
                                    st.write(f"  → Start: {dr['start_date']}, End: {dr['end_date']}")
                                    st.write(f"  → Years: {dr['years']}")
                            
                            year_debug_df = pd.DataFrame(debug_info['year_determination_debug'])
                            st.dataframe(year_debug_df, use_container_width=True)
                            
                            # Summary statistics
                            st.write("**Summary:**")
                            year_markers = [d for d in debug_info['year_determination_debug'] if d['type'] == 'YEAR_MARKER_DETECTED']
                            year_markers_rejected = [d for d in debug_info['year_determination_debug'] if d['type'] == 'YEAR_MARKER_REJECTED']
                            transaction_dates = [d for d in debug_info['year_determination_debug'] if d['type'] == 'TRANSACTION_DATE']
                            chrono_adjustments = [d for d in transaction_dates if 'chronological_adjustment' in d and 'incremented' in d.get('chronological_adjustment', '')]
                            date_range_usage = [d for d in transaction_dates if 'date_range_used' in d]
                            
                            col1, col2, col3, col4, col5 = st.columns(5)
                            col1.metric("Year Markers", len(year_markers))
                            col2.metric("Rejected Markers", len(year_markers_rejected))
                            col3.metric("Transaction Dates", len(transaction_dates))
                            col4.metric("Chronological Adj.", len(chrono_adjustments))
                            col5.metric("Date Range Usage", len(date_range_usage))
                            
                            # Show year markers
                            if year_markers:
                                st.write("**Year Markers Found:**")
                                for marker in year_markers:
                                    st.write(f"  Page {marker['page']}: '{marker['line']}' → Year {marker['detected_year']}")
                            
                            # Show rejected year markers
                            if year_markers_rejected:
                                st.write("**Rejected Year Markers:**")
                                for marker in year_markers_rejected:
                                    st.write(f"  Page {marker['page']}: '{marker['line']}' → Year {marker['detected_year']} ({marker['reason']})")
                            
                            # Show chronological adjustments
                            if chrono_adjustments:
                                st.write("**Chronological Year Adjustments:**")
                                for adj in chrono_adjustments:
                                    st.write(f"  Page {adj['page']}: '{adj['line']}' → {adj['chronological_adjustment']}")
                                    st.write(f"    Before: {adj['current_date_before_chrono']}, After: {adj['date_str_after_chrono']}")
                            
                            # Show date range usage
                            if date_range_usage:
                                st.write("**Date Range Usage:**")
                                for dr in date_range_usage:
                                    st.write(f"  Page {dr['page']}: '{dr['line']}' → {dr['date_range_used']}")
                                    st.write(f"    Final date: {dr['final_date_str']}")
                        
                        # Show date parsing debug
                        if debug_info.get('date_parsing_debug'):
                            st.write("---")
                            st.subheader("📅 Date Parsing Debug")
                            
                            date_parsing_df = pd.DataFrame(debug_info['date_parsing_debug'])
                            st.dataframe(date_parsing_df, use_container_width=True)
                            
                            # Show lines where dates were detected
                            date_detected = [d for d in debug_info['date_parsing_debug'] if d['date_match']]
                            if date_detected:
                                st.write(f"**Dates Detected ({len(date_detected)} lines):**")
                                for d in date_detected:
                                    st.write(f"  Page {d['page']}: '{d['line']}'")
                                    st.write(f"    Raw: `{d['raw_date_str']}` → Parsed: `{d['parsed_date_str']}`")
                        
                        # Show reconciliation results
                        if debug_info.get('reconciliation'):
                            rec = debug_info['reconciliation']
                            st.write("---")
                            st.subheader("💰 Balance Reconciliation Check")
                            
                            if rec['starting_balance'] is not None:
                                st.write(f"**Starting Balance:** ${rec['starting_balance']:,.2f}")
                                st.write(f"**Final Running Balance:** ${rec['running_balance']:,.2f}")
                                
                                if rec['discrepancies']:
                                    st.error(f"⚠️ Found {len(rec['discrepancies'])} discrepancy(ies):")
                                    for disc in rec['discrepancies']:
                                        with st.expander(f"❌ Transaction #{disc['transaction_num']}: {disc['description']}... (Diff: ${disc['difference']:,.2f})"):
                                            st.write(f"**Previous Balance:** ${disc['previous_balance']:,.2f}")
                                            st.write(f"**Debit:** ${disc['debit']:,.2f}")
                                            st.write(f"**Credit:** ${disc['credit']:,.2f}")
                                            st.write(f"**Expected Balance:** ${disc['expected_balance']:,.2f}")
                                            st.write(f"**Actual Balance:** ${disc['actual_balance']:,.2f}")
                                            st.write(f"**Difference:** ${disc['difference']:,.2f}")
                                else:
                                    st.success("✅ All balances reconcile correctly!")
                            else:
                                st.warning("No starting balance detected - cannot perform reconciliation")
                    
                    # Create Excel with all content
                    uploaded_file.seek(0)
                    excel_file = create_bank_statement_excel_full(transactions_df, all_text_df, tables)
                    
                    # Download button
                    st.subheader("💾 Download")
                    file_name = uploaded_file.name.replace('.pdf', '_complete.xlsx')
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 Download Complete Excel",
                            data=excel_file,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    with col2:
                        if not transactions_df.empty:
                            # Also offer CSV of just transactions
                            csv = transactions_df.to_csv(index=False)
                            st.download_button(
                                label="📄 Download Transactions CSV",
                                data=csv,
                                file_name=file_name.replace('.xlsx', '_transactions.csv'),
                                mime="text/csv"
                            )
                        else:
                            # Regular extraction
                            tables = extract_tables_from_pdf(uploaded_file)
                    
                    # Extract text
                    uploaded_file.seek(0)
                    text_df = extract_text_to_dataframe(uploaded_file)
                    
                    # Create Excel file
                    uploaded_file.seek(0)
                    excel_file = create_excel_with_tables(tables, text_df)
                    
                    # Show preview
                    st.subheader("📊 Preview")
                    
                    if tables:
                        st.write(f"**Found {len(tables)} table(s)**")
                        for i, table_info in enumerate(tables[:3], 1):
                            with st.expander(f"Table {i} (Page {table_info['page']}) - Preview"):
                                st.dataframe(table_info['data'].head(10))
                    else:
                        st.info("No tables detected in the PDF.")
                    
                    if not text_df.empty:
                        with st.expander("Text Content Preview"):
                            st.dataframe(text_df.head(20))
                    
                    # Download button
                    st.subheader("💾 Download")
                    file_name = uploaded_file.name.replace('.pdf', '.xlsx')
                    
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                except Exception as e:
                    st.error(f"❌ Error processing PDF: {str(e)}")
                    st.info("💡 Tip: Make sure the PDF contains selectable text (not scanned images). For scanned PDFs, OCR would be required.")
    
    st.markdown("---")
    st.markdown("<p style='text-align: center; color: #666;'>Supports PDFs with tables and text content | Optimized for bank statements</p>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
